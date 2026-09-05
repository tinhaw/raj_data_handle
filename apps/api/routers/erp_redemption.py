"""Local redemption management APIs with every remote operation intentionally absent."""

from __future__ import annotations

import re
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from apps.api.dependencies import require_erp_permission
from packages.common.database import get_db_session
from packages.domain.schemas.erp_redemption import (
    ErpRedemptionBatchCreateRequest,
    ErpRedemptionBatchDetailResponse,
    ErpRedemptionBatchResponse,
    ErpRedemptionCampaignCreateRequest,
    ErpRedemptionCampaignResponse,
    ErpRedemptionCodeImportRequest,
    ErpRedemptionLocalPublishRequest,
    ErpRedemptionTaskCreateRequest,
    ErpRedemptionTaskResponse,
)
from packages.domain.schemas.erp_redemption_remote import (
    ErpRedemptionRemoteExecutionResponse,
    ErpRedemptionRemotePlanRecoverRequest,
    ErpRedemptionRemotePlanResponse,
    ErpRedemptionRemotePlanWrite,
    ErpRedemptionRemotePublishPlanRequest,
    ErpRedemptionRemoteScheduleCancelRequest,
    ErpRedemptionTaskRemotePlanWrite,
)
from packages.domain.services.auth_service import AuthContext, write_audit
from packages.domain.services.erp_access_service import (
    ERP_PERMISSION_REDEMPTION_EXPORT,
    ERP_PERMISSION_REDEMPTION_GENERATE,
    ERP_PERMISSION_REDEMPTION_MANAGE,
    ERP_PERMISSION_REDEMPTION_VIEW,
)
from packages.domain.services.erp_redemption_remote_plan_service import (
    ErpRedemptionRemotePlanConflictError,
    ErpRedemptionRemotePlanError,
    ErpRedemptionRemotePlanNotFoundError,
    cancel_local_erp_redemption_publish_schedule,
    configure_erp_redemption_remote_plan,
    configure_erp_redemption_task_remote_plans,
    get_erp_redemption_remote_plan,
    list_due_erp_redemption_publish_plans,
    list_erp_redemption_remote_executions,
    plan_erp_redemption_remote_publish,
    recover_erp_redemption_remote_plan,
)
from packages.domain.services.erp_redemption_service import (
    ErpRedemptionConflictError,
    ErpRedemptionError,
    ErpRedemptionNotFoundError,
    create_erp_redemption_batch,
    create_erp_redemption_campaign,
    create_erp_redemption_task,
    get_erp_redemption_batch,
    get_erp_redemption_task,
    import_erp_redemption_codes,
    list_erp_redemption_batches,
    list_erp_redemption_campaigns,
    list_erp_redemption_tasks,
    publish_erp_redemption_batch_locally,
)

router = APIRouter(prefix="/erp/redemption", tags=["erp-redemption"])


def _api_error(exc: ErpRedemptionError) -> HTTPException:
    if isinstance(exc, ErpRedemptionNotFoundError):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    if isinstance(exc, ErpRedemptionConflictError):
        return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc))
    return HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


def _remote_api_error(exc: ErpRedemptionRemotePlanError) -> HTTPException:
    if isinstance(exc, ErpRedemptionRemotePlanNotFoundError):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    if isinstance(exc, ErpRedemptionRemotePlanConflictError):
        return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc))
    return HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


def _fit_columns(sheet: Worksheet) -> None:
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            28,
        )


def _append_issue_sheet(sheet: Worksheet, detail: ErpRedemptionBatchDetailResponse) -> None:
    sheet.append(
        [
            "领取日期",
            "充值窗口开始",
            "充值窗口结束",
            "档位",
            "充值门槛",
            "赠金",
            "最大奖金",
            "兑换码",
            "本地参考",
            "本地状态",
        ]
    )
    for issue in detail.issues:
        sheet.append(
            [
                issue.claim_date,
                issue.deposit_window_start,
                issue.deposit_window_end,
                issue.tier_name,
                issue.min_deposit_amount,
                issue.bonus_amount,
                issue.bonus_max_amount,
                issue.redemption_code,
                issue.local_reference,
                issue.workflow_status,
            ]
        )
    _fit_columns(sheet)


def _export_bytes(detail: ErpRedemptionBatchDetailResponse) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "兑换码"
    _append_issue_sheet(sheet, detail)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheet_title(order: int, source_name: str) -> str:
    safe_name = re.sub(r"[\\/*?:\[\]]", "-", source_name).strip() or "盘口"
    return f"{order:02d}-{safe_name}"[:31]


def _export_task_bytes(
    task: ErpRedemptionTaskResponse,
    details: list[ErpRedemptionBatchDetailResponse],
) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "任务概览"
    overview.append(["任务名称", task.task_name])
    overview.append(["领取日期", f"{task.claim_date_from} 至 {task.claim_date_to}"])
    overview.append(["回看天数", task.lookback_days])
    overview.append(["导出组标识", task.export_group_key])
    overview.append(["本地状态", task.status])
    overview.append(["已登记 / 总数", f"{task.imported_code_count} / {task.expected_code_count}"])
    overview.append([])
    overview.append(
        [
            "执行顺序",
            "盘口",
            "远端账号",
            "子任务 ID",
            "已登记",
            "总数",
            "本地状态",
        ]
    )
    for subtask in task.subtasks:
        overview.append(
            [
                subtask.execution_order,
                subtask.source_display_name,
                subtask.remote_account_name,
                subtask.batch_id,
                subtask.imported_code_count,
                subtask.expected_code_count,
                subtask.status,
            ]
        )
    _fit_columns(overview)

    for subtask, detail in zip(task.subtasks, details, strict=True):
        sheet = workbook.create_sheet(
            title=_sheet_title(subtask.execution_order, subtask.source_display_name)
        )
        _append_issue_sheet(sheet, detail)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@router.get("/campaigns", response_model=list[ErpRedemptionCampaignResponse])
async def get_campaigns(
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionCampaignResponse]:
    return await list_erp_redemption_campaigns(session)


@router.post(
    "/campaigns",
    response_model=ErpRedemptionCampaignResponse,
    status_code=status.HTTP_201_CREATED,
)
async def post_campaign(
    payload: ErpRedemptionCampaignCreateRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_MANAGE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionCampaignResponse:
    try:
        return await create_erp_redemption_campaign(
            session,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.get("/tasks", response_model=list[ErpRedemptionTaskResponse])
async def get_tasks(
    campaign_id: str | None = None,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionTaskResponse]:
    try:
        return await list_erp_redemption_tasks(session, campaign_id=campaign_id)
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.get("/tasks/{task_id}", response_model=ErpRedemptionTaskResponse)
async def get_task(
    task_id: str,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionTaskResponse:
    try:
        return await get_erp_redemption_task(session, task_id=task_id)
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.get("/tasks/{task_id}/export")
async def export_task(
    task_id: str,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_EXPORT)),
    session: AsyncSession = Depends(get_db_session),
) -> Response:
    try:
        task = await get_erp_redemption_task(session, task_id=task_id)
        details = [
            await get_erp_redemption_batch(session, batch_id=subtask.batch_id)
            for subtask in task.subtasks
        ]
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc
    issue_count = sum(len(detail.issues) for detail in details)
    await write_audit(
        session,
        action="erp_redemption_task.export",
        actor_user_id=auth.user.id,
        target_type="erp_redemption_task",
        target_id=task_id,
        metadata={
            "subtask_count": len(task.subtasks),
            "issue_count": issue_count,
            "export_group_key": task.export_group_key,
        },
    )
    await session.commit()
    return Response(
        content=_export_task_bytes(task, details),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="erp-redemption-task-{task_id}.xlsx"'
            )
        },
    )


@router.get("/campaigns/{campaign_id}/batches", response_model=list[ErpRedemptionBatchResponse])
async def get_campaign_batches(
    campaign_id: str,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionBatchResponse]:
    try:
        return await list_erp_redemption_batches(session, campaign_id=campaign_id)
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.post(
    "/batches",
    response_model=ErpRedemptionBatchDetailResponse,
    status_code=status.HTTP_201_CREATED,
)
async def post_batch(
    payload: ErpRedemptionBatchCreateRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionBatchDetailResponse:
    try:
        return await create_erp_redemption_batch(
            session,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.post(
    "/tasks",
    response_model=ErpRedemptionTaskResponse,
    status_code=status.HTTP_201_CREATED,
)
async def post_task(
    payload: ErpRedemptionTaskCreateRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionTaskResponse:
    try:
        return await create_erp_redemption_task(
            session,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.put(
    "/tasks/{task_id}/remote-plans",
    response_model=list[ErpRedemptionRemotePlanResponse],
)
async def put_task_remote_plans(
    task_id: str,
    payload: ErpRedemptionTaskRemotePlanWrite,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionRemotePlanResponse]:
    try:
        return await configure_erp_redemption_task_remote_plans(
            session,
            task_id=task_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.get("/batches/{batch_id}", response_model=ErpRedemptionBatchDetailResponse)
async def get_batch(
    batch_id: str,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionBatchDetailResponse:
    try:
        return await get_erp_redemption_batch(session, batch_id=batch_id)
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.get(
    "/batches/{batch_id}/remote-plan",
    response_model=ErpRedemptionRemotePlanResponse | None,
)
async def get_remote_plan(
    batch_id: str,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionRemotePlanResponse | None:
    try:
        return await get_erp_redemption_remote_plan(session, batch_id=batch_id)
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.put(
    "/batches/{batch_id}/remote-plan",
    response_model=ErpRedemptionRemotePlanResponse,
)
async def put_remote_plan(
    batch_id: str,
    payload: ErpRedemptionRemotePlanWrite,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionRemotePlanResponse:
    try:
        return await configure_erp_redemption_remote_plan(
            session,
            batch_id=batch_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.post(
    "/batches/{batch_id}/remote-plan/publish",
    response_model=ErpRedemptionRemotePlanResponse,
)
async def post_remote_publish_plan(
    batch_id: str,
    payload: ErpRedemptionRemotePublishPlanRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionRemotePlanResponse:
    try:
        return await plan_erp_redemption_remote_publish(
            session,
            batch_id=batch_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.post(
    "/batches/{batch_id}/remote-plan/publish/cancel-local",
    response_model=ErpRedemptionRemotePlanResponse,
)
async def post_cancel_local_publish_schedule(
    batch_id: str,
    payload: ErpRedemptionRemoteScheduleCancelRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionRemotePlanResponse:
    try:
        return await cancel_local_erp_redemption_publish_schedule(
            session,
            batch_id=batch_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.post(
    "/batches/{batch_id}/remote-plan/recover",
    response_model=ErpRedemptionRemotePlanResponse,
)
async def post_recover_remote_plan(
    batch_id: str,
    payload: ErpRedemptionRemotePlanRecoverRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionRemotePlanResponse:
    try:
        return await recover_erp_redemption_remote_plan(
            session,
            batch_id=batch_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.get(
    "/batches/{batch_id}/remote-executions",
    response_model=list[ErpRedemptionRemoteExecutionResponse],
)
async def get_remote_executions(
    batch_id: str,
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_VIEW)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionRemoteExecutionResponse]:
    try:
        return await list_erp_redemption_remote_executions(session, batch_id=batch_id)
    except ErpRedemptionRemotePlanError as exc:
        raise _remote_api_error(exc) from exc


@router.get(
    "/remote-plans/due",
    response_model=list[ErpRedemptionRemotePlanResponse],
)
async def get_due_remote_publish_plans(
    _: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> list[ErpRedemptionRemotePlanResponse]:
    return await list_due_erp_redemption_publish_plans(session)


@router.post("/batches/{batch_id}/codes", response_model=ErpRedemptionBatchDetailResponse)
async def post_codes(
    batch_id: str,
    payload: ErpRedemptionCodeImportRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionBatchDetailResponse:
    try:
        return await import_erp_redemption_codes(
            session,
            batch_id=batch_id,
            request=payload,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.post("/batches/{batch_id}/publish-local", response_model=ErpRedemptionBatchDetailResponse)
async def post_publish_local(
    batch_id: str,
    payload: ErpRedemptionLocalPublishRequest,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_GENERATE)),
    session: AsyncSession = Depends(get_db_session),
) -> ErpRedemptionBatchDetailResponse:
    try:
        return await publish_erp_redemption_batch_locally(
            session,
            batch_id=batch_id,
            row_version=payload.row_version,
            actor_user_id=auth.user.id,
        )
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc


@router.get("/batches/{batch_id}/export")
async def export_batch(
    batch_id: str,
    auth: AuthContext = Depends(require_erp_permission(ERP_PERMISSION_REDEMPTION_EXPORT)),
    session: AsyncSession = Depends(get_db_session),
) -> Response:
    try:
        detail = await get_erp_redemption_batch(session, batch_id=batch_id)
    except ErpRedemptionError as exc:
        raise _api_error(exc) from exc
    await write_audit(
        session,
        action="erp_redemption_batch.export",
        actor_user_id=auth.user.id,
        target_type="erp_redemption_batch",
        target_id=batch_id,
        metadata={"issue_count": len(detail.issues)},
    )
    await session.commit()
    return Response(
        content=_export_bytes(detail),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="erp-redemption-{batch_id}.xlsx"'},
    )
