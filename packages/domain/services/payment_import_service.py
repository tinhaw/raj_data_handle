from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook


class PaymentImportError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class PaymentOrderGroup:
    order_group_id: str
    merchant_order_no: str | None
    platform_order_no: str | None
    amount: Decimal | None
    currency: str
    payment_status_raw: str | None
    payment_status_group: str
    payment_time: datetime | None
    source_sheet: str
    source_row_numbers: list[int]
    duplicate_count: int
    preliminary_result_status: str | None
    platform_key: str


@dataclass(frozen=True, slots=True)
class PaymentImportResult:
    groups: list[PaymentOrderGroup]
    source_rows: int
    included_rows: int
    excluded_outside_window: int


def _order_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = str(value).strip()
    return normalized or None


def _decimal(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


def _parse_datetime(value: object, timezone_name: str) -> datetime | None:
    if value is None or str(value).strip() in {"", "0"}:
        return None
    try:
        timezone = ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise PaymentImportError("支付平台时区不是有效 IANA 时区。") from exc
    if isinstance(value, datetime):
        parsed = value
    else:
        raw = str(value).strip()
        parsed = None
        for parser in (
            datetime.fromisoformat,
            lambda item: datetime.strptime(item, "%Y/%m/%d %H:%M:%S"),
            lambda item: datetime.strptime(item, "%d-%m-%Y %H:%M:%S"),
        ):
            try:
                parsed = parser(raw)
                break
            except ValueError:
                continue
        if parsed is None:
            return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone)
    return parsed.astimezone(timezone)


def _window_datetime(value: str, timezone_name: str) -> datetime:
    parsed = _parse_datetime(value, timezone_name)
    if parsed is None:
        raise PaymentImportError("用户确认的时间范围格式无效。")
    return parsed


def _xlsx_rows(
    path: Path,
    *,
    sheet_name: str,
    header_row: int,
) -> Iterable[tuple[int, dict[str, object]]]:
    file_object = path.open("rb")
    workbook = load_workbook(file_object, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise PaymentImportError("支付文件中找不到已确认的工作表。")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers: list[str] | None = None
        for row_number, values in enumerate(rows, start=1):
            if row_number < header_row:
                continue
            if row_number == header_row:
                headers = [str(value).strip() if value is not None else "" for value in values]
                continue
            assert headers is not None
            if not any(value is not None and str(value).strip() for value in values):
                continue
            yield row_number, dict(zip(headers, values, strict=False))
    finally:
        workbook.close()
        file_object.close()


def _csv_rows(
    path: Path,
    *,
    header_row: int,
) -> Iterable[tuple[int, dict[str, object]]]:
    raw = path.read_bytes()
    decoded: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise PaymentImportError("CSV 编码无法识别。")
    rows = list(csv.reader(io.StringIO(decoded)))
    if len(rows) < header_row:
        raise PaymentImportError("CSV 中找不到已确认的表头行。")
    headers = [str(value).strip() for value in rows[header_row - 1]]
    for row_number, values in enumerate(rows[header_row:], start=header_row + 1):
        if not any(str(value).strip() for value in values):
            continue
        yield row_number, dict(zip(headers, values, strict=False))


def import_payment_orders(
    path: Path,
    *,
    file_suffix: str | None = None,
    platform_key: str,
    source_sheet: str,
    header_row: int,
    column_mapping: dict[str, Any],
    success_status_values: list[str],
    payment_time_field: str,
    payment_timezone: str,
    window_start: str,
    window_end: str,
    currency: str,
) -> PaymentImportResult:
    suffix = (file_suffix or path.suffix).lower()
    if suffix == ".xlsx":
        rows = _xlsx_rows(path, sheet_name=source_sheet, header_row=header_row)
    elif suffix == ".csv":
        rows = _csv_rows(path, header_row=header_row)
    else:
        raise PaymentImportError("当前只支持 .xlsx 和 .csv 文件。")

    merchant_column = str(column_mapping.get("merchant_order_no") or "")
    platform_column = str(column_mapping.get("platform_order_no") or "")
    amount_column = str(column_mapping.get("amount") or "")
    status_column = str(column_mapping.get("payment_status") or "")
    if not merchant_column or not amount_column or not status_column:
        raise PaymentImportError("支付模板缺少订单号、金额或状态映射。")
    candidate_times = column_mapping.get("candidate_time_fields")
    if not isinstance(candidate_times, list) or payment_time_field not in candidate_times:
        raise PaymentImportError("支付时间列不属于当前模板允许的候选字段。")

    window_lower = _window_datetime(window_start, payment_timezone)
    window_upper = _window_datetime(window_end, payment_timezone)
    if window_lower > window_upper:
        raise PaymentImportError("支付时间范围开始时间不能晚于结束时间。")

    success_values = {str(value).strip() for value in success_status_values}
    source_rows = 0
    excluded = 0
    parsed_rows: list[dict[str, Any]] = []
    for row_number, row in rows:
        source_rows += 1
        payment_time = _parse_datetime(row.get(payment_time_field), payment_timezone)
        time_is_invalid = payment_time is None
        if payment_time is not None and not (window_lower <= payment_time <= window_upper):
            excluded += 1
            continue

        merchant_order_no = _order_text(row.get(merchant_column))
        platform_order_no = _order_text(row.get(platform_column)) if platform_column else None
        amount = _decimal(row.get(amount_column))
        payment_status_raw = _order_text(row.get(status_column))
        payment_status_group = (
            "success"
            if payment_status_raw in success_values
            else "non_success"
            if payment_status_raw
            else "unknown"
        )
        invalid = (
            not merchant_order_no or amount is None or payment_status_raw is None or time_is_invalid
        )
        parsed_rows.append(
            {
                "row_number": row_number,
                "merchant_order_no": merchant_order_no,
                "platform_order_no": platform_order_no,
                "amount": amount,
                "payment_status_raw": payment_status_raw,
                "payment_status_group": payment_status_group,
                "payment_time": payment_time,
                "invalid": invalid,
            }
        )

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in parsed_rows:
        group_key = row["merchant_order_no"] or f"invalid-row:{row['row_number']}"
        grouped.setdefault(group_key, []).append(row)

    groups: list[PaymentOrderGroup] = []
    for group_key, members in grouped.items():
        first = members[0]
        fingerprints = {
            (
                member["merchant_order_no"],
                member["platform_order_no"],
                str(member["amount"]) if member["amount"] is not None else None,
                member["payment_status_raw"],
                member["payment_time"].isoformat() if member["payment_time"] else None,
            )
            for member in members
        }
        preliminary_status = (
            "invalid_payment_row"
            if any(member["invalid"] for member in members)
            else "duplicate_payment_conflict"
            if len(fingerprints) > 1
            else None
        )
        group_id = hashlib.sha256(f"{platform_key}:{group_key}".encode()).hexdigest()
        groups.append(
            PaymentOrderGroup(
                order_group_id=group_id,
                merchant_order_no=first["merchant_order_no"],
                platform_order_no=first["platform_order_no"],
                amount=first["amount"],
                currency=currency,
                payment_status_raw=first["payment_status_raw"],
                payment_status_group=first["payment_status_group"],
                payment_time=first["payment_time"],
                source_sheet=source_sheet,
                source_row_numbers=[int(member["row_number"]) for member in members],
                duplicate_count=len(members),
                preliminary_result_status=preliminary_status,
                platform_key=platform_key,
            )
        )
    return PaymentImportResult(
        groups=groups,
        source_rows=source_rows,
        included_rows=len(parsed_rows),
        excluded_outside_window=excluded,
    )
