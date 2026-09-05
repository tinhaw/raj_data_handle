"""Preview-first local ERP daily-ledger imports with no remote side effects."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from packages.domain.models import (
    ErpDailyBalance,
    ErpImportJob,
    ErpImportJobRow,
    ErpOperatorLine,
)
from packages.domain.schemas.erp_balance import ErpDailyBalanceWriteRequest
from packages.domain.schemas.erp_import import (
    ErpImportCommitResponse,
    ErpImportJobResponse,
    ErpImportPreviewResponse,
    ErpImportRowResponse,
)
from packages.domain.services.auth_service import write_audit
from packages.domain.services.erp_balance_service import (
    ErpBalanceError,
    create_erp_daily_balance,
    preview_erp_daily_balance,
    update_erp_daily_balance,
)
from packages.domain.services.erp_operator_service import (
    ErpOperatorNotFoundError,
    get_erp_operator_line,
)

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000
CONFLICT_STRATEGIES = frozenset({"SKIP_EXISTING", "UPDATE_DRAFT", "REJECT_ON_CONFLICT"})


class ErpImportError(ValueError):
    pass


class ErpImportNotFoundError(ErpImportError):
    pass


class ErpImportConflictError(ErpImportError):
    pass


@dataclass(frozen=True, slots=True)
class _ParsedRow:
    source_sheet: str
    source_row: int
    source_json: dict[str, Any]
    request: ErpDailyBalanceWriteRequest | None
    error_code: str | None = None
    error_message: str | None = None


def _normalize_header(value: Any) -> str | None:
    source = str(value or "").strip()
    if not source:
        return None
    compact = "".join(char for char in source.upper() if char not in " _-－—\t\n")
    aliases = {
        "日期": "business_date",
        "业务日期": "business_date",
        "DATE": "business_date",
        "期初余额": "opening_balance",
        "期初结余": "opening_balance",
        "昨日结余": "opening_balance",
        "OPENING": "opening_balance",
        "转U": "transfer_amount",
        "转USDT": "transfer_amount",
        "TRANSFER": "transfer_amount",
        "消耗": "spend_amount",
        "SPEND": "spend_amount",
        "欺诈损失": "fraud_loss_amount",
        "FRAUD": "fraud_loss_amount",
        "欺诈承担": "fraud_deduction_source",
        "欺诈扣减来源": "fraud_deduction_source",
        "FRAUDSOURCE": "fraud_deduction_source",
        "汇损费率": "exchange_loss_rate",
        "汇损基数": "exchange_loss_basis",
        "汇损模式": "exchange_loss_mode",
        "汇损金额": "exchange_loss_amount",
        "服务费率": "service_fee_rate",
        "服务费基数": "service_fee_basis",
        "服务费模式": "service_fee_mode",
        "服务费金额": "service_fee_amount",
        "回流": "reflux_amount",
        "REFLUX": "reflux_amount",
        "退款": "refund_amount",
        "REFUND": "refund_amount",
        "其他扣减": "other_deduction_amount",
        "其他": "other_deduction_amount",
        "OTHER": "other_deduction_amount",
        "其他原因": "other_reason",
        "备注": "remark",
        "REMARK": "remark",
    }
    return aliases.get(compact)


def _date_value(value: Any, *, business_year: int | None) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        raise ErpImportError("业务日期不能为空。")
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "")
    try:
        return datetime.strptime(normalized, "%Y-%m-%d").date()
    except ValueError:
        pass
    parts = normalized.split("-")
    if business_year and len(parts) == 2:
        try:
            return date(business_year, int(parts[0]), int(parts[1]))
        except ValueError as exc:
            raise ErpImportError(f"日期格式不正确：{text}") from exc
    raise ErpImportError(f"日期格式不正确：{text}")


def _decimal_value(value: Any, *, field: str) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace(",", "").replace("，", "")
    text = text.replace("USDT", "").replace("USDC", "").strip()
    percentage = text.endswith("%")
    if percentage:
        text = text[:-1]
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ErpImportError(f"{field}金额格式不正确：{value}") from exc
    return result / Decimal("100") if percentage else result


def _text_value(value: Any) -> str | None:
    result = str(value or "").strip()
    return result or None


def _basis(value: Any) -> str | None:
    source = _text_value(value)
    if not source:
        return None
    lookup = {
        "转U": "TRANSFER",
        "有效转U": "EFFECTIVE_TRANSFER",
        "消耗": "SPEND",
        "手工": "MANUAL",
    }
    return lookup.get(source.replace(" ", ""), source.upper())


def _mode(value: Any) -> str | None:
    source = _text_value(value)
    if not source:
        return None
    return {"自动": "AUTO", "手工": "MANUAL"}.get(source, source.upper())


def _fraud_source(value: Any) -> str | None:
    source = _text_value(value)
    if not source:
        return None
    lookup = {"转U": "TRANSFER", "转账": "TRANSFER", "结余": "BALANCE"}
    return lookup.get(source.replace(" ", ""), source.upper())


def _request_from_values(
    values: dict[str, Any],
    *,
    operator_line_id: str,
    business_year: int | None,
) -> ErpDailyBalanceWriteRequest:
    opening = _decimal_value(values.get("opening_balance"), field="期初结余")
    exchange_amount = _decimal_value(values.get("exchange_loss_amount"), field="汇损")
    service_amount = _decimal_value(values.get("service_fee_amount"), field="服务费")
    exchange_mode = _mode(values.get("exchange_loss_mode"))
    service_mode = _mode(values.get("service_fee_mode"))
    return ErpDailyBalanceWriteRequest(
        operator_line_id=operator_line_id,
        business_date=_date_value(values.get("business_date"), business_year=business_year),
        opening_balance=opening,
        opening_mode="MANUAL" if opening is not None else "AUTO",
        transfer_amount=_decimal_value(values.get("transfer_amount"), field="转 U"),
        fraud_loss_amount=_decimal_value(values.get("fraud_loss_amount"), field="欺诈损失"),
        fraud_deduction_source=_fraud_source(values.get("fraud_deduction_source")),
        spend_amount=_decimal_value(values.get("spend_amount"), field="消耗"),
        exchange_loss_rate=_decimal_value(values.get("exchange_loss_rate"), field="汇损费率"),
        exchange_loss_basis=_basis(values.get("exchange_loss_basis")),
        exchange_loss_mode=exchange_mode or ("MANUAL" if exchange_amount is not None else None),
        exchange_loss_amount=exchange_amount,
        service_fee_rate=_decimal_value(values.get("service_fee_rate"), field="服务费率"),
        service_fee_basis=_basis(values.get("service_fee_basis")),
        service_fee_mode=service_mode or ("MANUAL" if service_amount is not None else None),
        service_fee_amount=service_amount,
        reflux_amount=_decimal_value(values.get("reflux_amount"), field="回流"),
        refund_amount=_decimal_value(values.get("refund_amount"), field="退款"),
        other_deduction_amount=_decimal_value(
            values.get("other_deduction_amount"),
            field="其他扣减",
        ),
        other_reason=_text_value(values.get("other_reason")),
        remark=_text_value(values.get("remark")),
        source_type="IMPORT",
    )


def _matrix_to_rows(
    matrix: list[list[Any]],
    *,
    source_sheet: str,
    operator_line_id: str,
    business_year: int | None,
) -> list[_ParsedRow]:
    nonempty = [
        (index, cells)
        for index, cells in enumerate(matrix, start=1)
        if any(str(cell or "").strip() for cell in cells)
    ]
    if not nonempty:
        raise ErpImportError("导入内容为空。")
    first_index, first_row = nonempty[0]
    headers = [_normalize_header(value) for value in first_row]
    has_header = "business_date" in headers
    if has_header:
        header_map = {key: index for index, key in enumerate(headers) if key}
        source_rows = nonempty[1:]
    else:
        header_map = {
            "business_date": 0,
            "opening_balance": 1,
            "transfer_amount": 2,
            "spend_amount": 3,
            "exchange_loss_amount": 4,
            "service_fee_amount": 5,
            "reflux_amount": 6,
            "refund_amount": 7,
            "other_deduction_amount": 8,
            "fraud_loss_amount": 9,
        }
        source_rows = nonempty
    if not source_rows:
        raise ErpImportError("导入内容缺少数据行。")
    if len(source_rows) > MAX_IMPORT_ROWS:
        raise ErpImportError(f"导入行数不能超过 {MAX_IMPORT_ROWS} 行。")
    parsed: list[_ParsedRow] = []
    for source_row, cells in source_rows:
        values = {
            key: cells[index] if index < len(cells) else None
            for key, index in header_map.items()
        }
        source_json = {key: value for key, value in values.items() if value not in (None, "")}
        try:
            request = _request_from_values(
                values,
                operator_line_id=operator_line_id,
                business_year=business_year,
            )
        except (ErpImportError, ValidationError) as exc:
            parsed.append(
                _ParsedRow(
                    source_sheet=source_sheet,
                    source_row=source_row,
                    source_json=source_json,
                    request=None,
                    error_code="IMPORT_ROW_INVALID",
                    error_message=str(exc),
                )
            )
        else:
            parsed.append(
                _ParsedRow(
                    source_sheet=source_sheet,
                    source_row=source_row,
                    source_json=source_json,
                    request=request,
                )
            )
    return parsed


def _paste_rows(
    text: str,
    *,
    operator_line_id: str,
    business_year: int | None,
) -> list[_ParsedRow]:
    matrix = [line.split("\t") for line in text.replace("\r\n", "\n").split("\n")]
    return _matrix_to_rows(
        matrix,
        source_sheet="粘贴",
        operator_line_id=operator_line_id,
        business_year=business_year,
    )


def _xlsx_rows(
    content: bytes,
    *,
    operator_line_id: str,
    business_year: int | None,
) -> list[_ParsedRow]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl emits several incompatible exception types.
        raise ErpImportError("无法读取 Excel 文件。") from exc
    parsed: list[_ParsedRow] = []
    for sheet in workbook.worksheets:
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not any(any(value is not None for value in row) for row in matrix):
            continue
        parsed.extend(
            _matrix_to_rows(
                matrix,
                source_sheet=sheet.title,
                operator_line_id=operator_line_id,
                business_year=business_year,
            )
        )
    if not parsed:
        raise ErpImportError("Excel 中没有可导入的数据行。")
    if len(parsed) > MAX_IMPORT_ROWS:
        raise ErpImportError(f"导入行数不能超过 {MAX_IMPORT_ROWS} 行。")
    return parsed


def _strategy(value: str) -> str:
    normalized = value.strip().upper()
    if normalized not in CONFLICT_STRATEGIES:
        raise ErpImportError("导入冲突策略不合法。")
    return normalized


def _job_response(job: ErpImportJob) -> ErpImportJobResponse:
    data = ErpImportJobResponse.model_validate(job).model_dump()
    data.update(
        source_available=bool(job.source_storage_key),
        error_report_available=job.error_rows > 0,
        source_size_bytes=job.source_size_bytes,
    )
    return ErpImportJobResponse.model_validate(data)


def _row_response(row: ErpImportJobRow) -> ErpImportRowResponse:
    normalized = (
        ErpDailyBalanceWriteRequest.model_validate(row.normalized_json)
        if row.normalized_json
        else None
    )
    return ErpImportRowResponse(
        id=row.id,
        source_sheet=row.source_sheet,
        source_row=row.source_row,
        source_json=row.source_json,
        operator_line_id=row.operator_line_id,
        business_date=row.business_date,
        severity=row.severity,
        error_code=row.error_code,
        error_message=row.error_message,
        action=row.action,
        target_daily_balance_id=row.target_daily_balance_id,
        normalized=normalized,
    )


async def _require_line(session: AsyncSession, *, line_id: str) -> None:
    try:
        line = await get_erp_operator_line(session, line_id=line_id)
    except ErpOperatorNotFoundError as exc:
        raise ErpImportNotFoundError(str(exc)) from exc
    if line.status != "ACTIVE":
        raise ErpImportConflictError("已停用的投放线不能导入。")


async def _persist_preview(
    session: AsyncSession,
    *,
    source_type: str,
    original_filename: str | None,
    file_sha256: str | None,
    source_storage_key: str | None,
    source_size_bytes: int | None,
    strategy: str,
    parsed_rows: list[_ParsedRow],
    actor_user_id: int,
) -> ErpImportPreviewResponse:
    job = ErpImportJob(
        source_type=source_type,
        original_filename=original_filename,
        file_sha256=file_sha256,
        source_storage_key=source_storage_key,
        source_size_bytes=source_size_bytes,
        status="PREVIEW_READY",
        conflict_strategy=strategy,
        created_by=actor_user_id,
    )
    session.add(job)
    await session.flush()
    seen: set[tuple[str, date]] = set()
    rows: list[ErpImportJobRow] = []
    valid = warning = errors = 0
    for parsed in parsed_rows:
        row = ErpImportJobRow(
            import_job_id=job.id,
            source_sheet=parsed.source_sheet,
            source_row=parsed.source_row,
            source_json=parsed.source_json,
        )
        if parsed.request is None:
            row.severity = "ERROR"
            row.error_code = parsed.error_code
            row.error_message = parsed.error_message
            errors += 1
        else:
            request = parsed.request
            row.operator_line_id = request.operator_line_id
            row.business_date = request.business_date
            row.normalized_json = request.model_dump(mode="json", by_alias=False)
            key = (request.operator_line_id, request.business_date)
            if key in seen:
                row.severity = "WARNING"
                row.error_code = "DUPLICATE_IN_FILE"
                row.error_message = "同一批次存在相同投放线和业务日期，后续行会跳过。"
                row.action = "SKIP"
                warning += 1
            else:
                seen.add(key)
                try:
                    await preview_erp_daily_balance(session, request=request)
                except ErpBalanceError as exc:
                    row.severity = "ERROR"
                    row.error_code = "IMPORT_ROW_INVALID"
                    row.error_message = str(exc)
                    errors += 1
                else:
                    existing = await session.scalar(
                        select(ErpDailyBalance).where(
                            ErpDailyBalance.operator_line_id == request.operator_line_id,
                            ErpDailyBalance.business_date == request.business_date,
                        )
                    )
                    if existing:
                        row.severity = "WARNING"
                        row.error_code = "DUPLICATE_RECORD"
                        row.error_message = "系统中已有该投放线和业务日期的日结记录。"
                        row.action = "SKIP"
                        row.preview_daily_balance_id = existing.id
                        row.preview_row_version = existing.row_version
                        warning += 1
                    else:
                        row.severity = "OK"
                        row.action = "CREATE"
                        valid += 1
        rows.append(row)
        session.add(row)
    job.total_rows = len(rows)
    job.valid_rows = valid
    job.warning_rows = warning
    job.error_rows = errors
    await session.flush()
    result = ErpImportPreviewResponse(
        job=_job_response(job),
        rows=[_row_response(row) for row in rows],
    )
    await write_audit(
        session,
        action="erp_import.preview",
        actor_user_id=actor_user_id,
        target_type="erp_import_job",
        target_id=job.id,
        metadata={"source_type": source_type, "total_rows": job.total_rows, "error_rows": errors},
    )
    await session.commit()
    return result


async def preview_erp_paste_import(
    session: AsyncSession,
    *,
    text: str,
    operator_line_id: str,
    conflict_strategy: str,
    business_year: int | None,
    actor_user_id: int,
) -> ErpImportPreviewResponse:
    await _require_line(session, line_id=operator_line_id)
    parsed = _paste_rows(text, operator_line_id=operator_line_id, business_year=business_year)
    return await _persist_preview(
        session,
        source_type="PASTE",
        original_filename=None,
        file_sha256=None,
        source_storage_key=None,
        source_size_bytes=None,
        strategy=_strategy(conflict_strategy),
        parsed_rows=parsed,
        actor_user_id=actor_user_id,
    )


async def preview_erp_excel_import(
    session: AsyncSession,
    *,
    content: bytes,
    original_filename: str,
    operator_line_id: str,
    conflict_strategy: str,
    business_year: int | None,
    actor_user_id: int,
    source_storage_key: str | None = None,
    source_size_bytes: int | None = None,
) -> ErpImportPreviewResponse:
    if len(content) > MAX_IMPORT_BYTES:
        raise ErpImportError("Excel 文件超过 10 MB 限制。")
    if not original_filename.lower().endswith(".xlsx"):
        raise ErpImportError("仅支持 .xlsx 文件。")
    await _require_line(session, line_id=operator_line_id)
    parsed = _xlsx_rows(content, operator_line_id=operator_line_id, business_year=business_year)
    return await _persist_preview(
        session,
        source_type="XLSX",
        original_filename=original_filename,
        file_sha256=sha256(content).hexdigest(),
        source_storage_key=source_storage_key,
        source_size_bytes=source_size_bytes,
        strategy=_strategy(conflict_strategy),
        parsed_rows=parsed,
        actor_user_id=actor_user_id,
    )


async def get_erp_import_job(session: AsyncSession, *, job_id: str) -> ErpImportPreviewResponse:
    job = await session.get(ErpImportJob, job_id)
    if job is None:
        raise ErpImportNotFoundError("导入批次不存在。")
    rows = list(
        (
            await session.scalars(
                select(ErpImportJobRow)
                .where(ErpImportJobRow.import_job_id == job.id)
                .order_by(ErpImportJobRow.source_sheet.asc(), ErpImportJobRow.source_row.asc())
            )
        ).all()
    )
    return ErpImportPreviewResponse(
        job=_job_response(job),
        rows=[_row_response(row) for row in rows],
    )


async def list_erp_import_jobs(
    session: AsyncSession, *, operator_ids: list[str] | None = None
) -> list[ErpImportJobResponse]:
    statement = select(ErpImportJob)
    if operator_ids is not None:
        line_ids = select(ErpOperatorLine.id).where(
            ErpOperatorLine.operator_id.in_(operator_ids)
        )
        statement = statement.where(
            ErpImportJob.id.in_(
                select(ErpImportJobRow.import_job_id).where(
                    ErpImportJobRow.operator_line_id.in_(line_ids)
                )
            )
        )
    rows = await session.scalars(statement.order_by(ErpImportJob.created_at.desc()))
    return [_job_response(row) for row in rows]


async def commit_erp_import_job(
    session: AsyncSession,
    *,
    job_id: str,
    conflict_strategy: str | None,
    actor_user_id: int,
) -> ErpImportCommitResponse:
    job = await session.get(ErpImportJob, job_id)
    if job is None:
        raise ErpImportNotFoundError("导入批次不存在。")
    if job.status != "PREVIEW_READY":
        raise ErpImportConflictError("该导入批次不能重复提交。")
    if job.error_rows:
        raise ErpImportError("导入预览仍有错误，不能提交。")
    strategy = _strategy(conflict_strategy or job.conflict_strategy)
    rows = list(
        (
            await session.scalars(
                select(ErpImportJobRow)
                .where(ErpImportJobRow.import_job_id == job.id)
                .order_by(ErpImportJobRow.source_sheet.asc(), ErpImportJobRow.source_row.asc())
            )
        ).all()
    )
    for row in rows:
        if not row.normalized_json or row.error_code == "DUPLICATE_IN_FILE":
            continue
        request = ErpDailyBalanceWriteRequest.model_validate(row.normalized_json)
        existing = await session.scalar(
            select(ErpDailyBalance).where(
                ErpDailyBalance.operator_line_id == request.operator_line_id,
                ErpDailyBalance.business_date == request.business_date,
            )
        )
        if strategy == "REJECT_ON_CONFLICT" and existing is not None:
            raise ErpImportConflictError("存在重复日结记录，请选择跳过或更新草稿。")
        if strategy == "UPDATE_DRAFT" and existing is not None:
            if (
                existing.status != "DRAFT"
                or existing.id != row.preview_daily_balance_id
                or existing.row_version != row.preview_row_version
            ):
                raise ErpImportConflictError("预览后目标草稿已变化，请重新预览后提交。")
    created = updated = skipped = 0
    for row in rows:
        if not row.normalized_json or row.error_code == "DUPLICATE_IN_FILE":
            row.action = "SKIPPED"
            skipped += 1
            continue
        request = ErpDailyBalanceWriteRequest.model_validate(row.normalized_json)
        existing = await session.scalar(
            select(ErpDailyBalance).where(
                ErpDailyBalance.operator_line_id == request.operator_line_id,
                ErpDailyBalance.business_date == request.business_date,
            )
        )
        if existing is None:
            saved = await create_erp_daily_balance(
                session,
                request=request,
                actor_user_id=actor_user_id,
                commit=False,
            )
            row.target_daily_balance_id = saved.id
            row.action = "CREATED"
            created += 1
        elif strategy == "UPDATE_DRAFT":
            saved = await update_erp_daily_balance(
                session,
                balance_id=existing.id,
                request=request.model_copy(update={"row_version": existing.row_version}),
                actor_user_id=actor_user_id,
                commit=False,
            )
            row.target_daily_balance_id = saved.id
            row.action = "UPDATED"
            updated += 1
        else:
            row.target_daily_balance_id = existing.id
            row.action = "SKIPPED"
            skipped += 1
    job.status = "SUCCEEDED"
    job.conflict_strategy = strategy
    job.committed_by = actor_user_id
    job.committed_at = datetime.now(UTC)
    await session.flush()
    result = ErpImportCommitResponse(
        job=_job_response(job),
        created=created,
        updated=updated,
        skipped=skipped,
    )
    await write_audit(
        session,
        action="erp_import.commit",
        actor_user_id=actor_user_id,
        target_type="erp_import_job",
        target_id=job.id,
        metadata={"created": created, "updated": updated, "skipped": skipped},
    )
    await session.commit()
    return result
