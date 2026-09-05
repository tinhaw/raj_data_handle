from __future__ import annotations

from collections.abc import Awaitable, Callable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

from packages.common.totp import generate_totp
from packages.domain.services.remote_account_session_service import (
    RemoteAccountSession,
    RemoteSessionError,
    response_requires_relogin,
)

LOGIN_PATH = "/api/system/login"
CHARGE_ORDER_INDEX_PATH = "/api/operate/chargeOrder/index"
CHARGE_CHANNEL_PATH = "/api/operate/chargeOrder/payChannel"
CHARGE_ORDER_EXPORT_PATH = "/api/operate/chargeOrder/export"
EXPORT_TASK_SAVE_PATH = "/api/operate/exportTask/save"
WITHDRAW_ORDER_INDEX_PATH = "/api/operate/withdrawOrder/index"
DATA_DICTIONARY_PATH = "/api/system/dataDict/list"
SPIN_ORDER_INDEX_PATH = "/api/operate/spinOrder/index"
PLAYER_INFO_LIST_PATH = "/api/operate/playerInfoList/index"
USER_SOURCE_CHANNEL_DICTIONARY_PATH = "/api/stat/userPayLtvLog/channel"
WITHDRAW_STATUS_DICTIONARY_PATH = DATA_DICTIONARY_PATH
REMOTE_SUCCESS_STATUS = 1
REMOTE_GET_PATHS = {
    CHARGE_ORDER_INDEX_PATH,
    CHARGE_CHANNEL_PATH,
    DATA_DICTIONARY_PATH,
    SPIN_ORDER_INDEX_PATH,
    PLAYER_INFO_LIST_PATH,
    USER_SOURCE_CHANNEL_DICTIONARY_PATH,
}
REMOTE_POST_PATHS = {
    WITHDRAW_ORDER_INDEX_PATH,
    CHARGE_ORDER_EXPORT_PATH,
    EXPORT_TASK_SAVE_PATH,
}
AUTH_FAILURE_STATUSES = {401, 403, 419, 440}
MAX_CHARGE_PAGES_PER_CHANNEL = 200
MAX_CHARGE_EXPORT_BYTES = 128 * 1024 * 1024
# Excel generation on a remote back office can take substantially longer than
# a normal JSON query.  Keep connection establishment bounded, while allowing
# a completed download enough time to arrive.
REMOTE_CONNECT_TIMEOUT_SECONDS = 10.0
REMOTE_REQUEST_TIMEOUT_SECONDS = 180.0

CHARGE_EXPORT_COLUMNS = (
    "订单id",
    "用户uid",
    "我方订单号",
    "充值商品id",
    "商品名称",
    "支付渠道名称",
    "支付渠道",
    "支付方式",
    "三方支付订单号",
    "支付金额",
    "发放金额",
    "赠送金额",
    "订单状态",
    "创建时间",
    "支付时间",
    "完成时间",
    "是否首充",
    "用户渠道",
)
CHARGE_EXPORT_STATUS_CODES = {
    "已失效": "-1",
    "未支付": "0",
    "待支付": "0",
    "已支付": "1",
    "已退款": "2",
}
CHARGE_EXPORT_EXCLUDED_STATUSES = frozenset({"测试拉单"})


class RemoteChargeError(RuntimeError):
    pass


class RemoteAuthenticationError(RemoteChargeError):
    pass


class RemoteResponseError(RemoteChargeError):
    pass


@dataclass(frozen=True, slots=True)
class ExactSearchResult:
    orders: list[dict[str, Any]]
    complete: bool


@dataclass(frozen=True, slots=True)
class ChargeFetchResult:
    orders: list[dict[str, Any]]
    fetched_pages: int
    remote_total: int
    complete: bool

    def __iter__(self):
        """Keep the existing reconciliation-client tuple contract intact."""

        yield self.orders
        yield self.fetched_pages


def _text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized if normalized and normalized != "-" else None


def _amount(value: object) -> str | None:
    text = _text(value)
    if text is None:
        return None
    try:
        return format(Decimal(text), "f")
    except (InvalidOperation, ValueError):
        return None


def normalize_charge_order(item: dict[str, Any]) -> dict[str, Any]:
    """Keep only the approved recharge-monitor fields from a remote item."""

    return {
        "id": _text(item.get("id")) or "",
        "uid": _text(item.get("uid")) or "",
        "order_num": _text(item.get("order_num")),
        "out_trade_no": _text(item.get("out_trade_no")),
        "pay_method": _text(item.get("pay_method") or item.get("_remote_channel_code")),
        "pay_channel_name": _text(item.get("pay_channel_name")),
        "amount": _amount(item.get("amount")),
        "balance": _amount(item.get("balance")),
        "extra": _amount(item.get("extra")),
        "status": _text(item.get("status")) or "",
        "create_time": _text(item.get("create_time")),
        "pay_time": _text(item.get("pay_time")),
        "update_time": _text(item.get("update_time")),
        "first_pay": _text(item.get("first_pay")),
        "notified": _text(item.get("notified")),
        "charge_type": _text(item.get("charge_type")),
        "channel": _text(item.get("channel")),
        "fill_order_id": _text(item.get("fill_order_id")),
        "fill_order_num": _text(item.get("fill_order_num")),
        "fill_order_admin": _text(item.get("fill_order_admin")),
    }


def _excel_text(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value)


def _export_status(value: object) -> str:
    text = _excel_text(value) or ""
    if text in {"-1", "0", "1", "2"}:
        return text
    code = CHARGE_EXPORT_STATUS_CODES.get(text)
    if code is None:
        raise RemoteResponseError("充值订单导出表格包含未识别的订单状态。")
    return code


def normalize_charge_order_export_row(item: dict[str, object]) -> dict[str, Any]:
    """Map the approved recharge-export columns into the local cache schema."""

    return {
        "id": _excel_text(item.get("订单id")) or "",
        "uid": _excel_text(item.get("用户uid")) or "",
        "order_num": _excel_text(item.get("我方订单号")),
        "charge_product_id": _excel_text(item.get("充值商品id")),
        "product_name": _excel_text(item.get("商品名称")),
        "pay_channel_name": _excel_text(item.get("支付渠道名称")),
        "pay_method": _excel_text(item.get("支付渠道")),
        "pay_type": _excel_text(item.get("支付方式")),
        "out_trade_no": _excel_text(item.get("三方支付订单号")),
        "amount": _amount(_excel_text(item.get("支付金额"))),
        "balance": _amount(_excel_text(item.get("发放金额"))),
        "extra": _amount(_excel_text(item.get("赠送金额"))),
        "status": _export_status(item.get("订单状态")),
        "create_time": _excel_text(item.get("创建时间")),
        "pay_time": _excel_text(item.get("支付时间")),
        "update_time": _excel_text(item.get("完成时间")),
        "first_pay": _excel_text(item.get("是否首充")),
        "channel": _excel_text(item.get("用户渠道")),
    }


def parse_charge_order_export(content: bytes) -> list[dict[str, Any]]:
    """Read the daily recharge export without persisting the workbook itself."""

    if not content or len(content) > MAX_CHARGE_EXPORT_BYTES:
        raise RemoteResponseError("充值订单导出文件为空或超过大小限制。")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception as exc:
        raise RemoteResponseError("远端充值订单导出文件不是有效 Excel 表格。") from exc
    if header_row is None:
        raise RemoteResponseError("远端充值订单导出表格缺少表头。")
    headers = [_excel_text(value) or "" for value in header_row]
    missing = [column for column in CHARGE_EXPORT_COLUMNS if column not in headers]
    if missing:
        raise RemoteResponseError("远端充值订单导出表格缺少必要列。")
    indexes = {header: index for index, header in enumerate(headers)}
    orders: dict[str, dict[str, Any]] = {}
    try:
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                column: values[indexes[column]] if indexes[column] < len(values) else None
                for column in CHARGE_EXPORT_COLUMNS
            }
            if not any(value not in (None, "") for value in row.values()):
                continue
            if _excel_text(row.get("订单状态")) in CHARGE_EXPORT_EXCLUDED_STATUSES:
                continue
            order = normalize_charge_order_export_row(row)
            if not order["id"]:
                raise RemoteResponseError("充值订单导出表格包含缺少订单id的行。")
            orders[order["id"]] = order
    finally:
        workbook.close()
    return list(orders.values())


def _extract_token(payload: object) -> str | None:
    token_keys = {"token", "jwt", "access_token", "accessToken"}
    if isinstance(payload, dict):
        for key in token_keys:
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                token = value.strip()
                if token.lower().startswith("bearer "):
                    return token.split(None, 1)[1].strip()
                return token
        for value in payload.values():
            token = _extract_token(value)
            if token:
                return token
    elif isinstance(payload, list):
        for value in payload:
            token = _extract_token(value)
            if token:
                return token
    return None


def _response_data(payload: object) -> Any:
    if not isinstance(payload, dict):
        raise RemoteResponseError("远端响应不是 JSON 对象。")
    if payload.get("success") is False:
        raise RemoteResponseError("远端接口返回失败状态。")
    code = payload.get("code")
    if code not in (None, 0, 200, "200"):
        raise RemoteResponseError("远端接口返回非成功业务码。")
    return payload.get("data")


class RajAdminChargeClient:
    def __init__(
        self,
        *,
        base_url: str,
        username: str,
        password: str,
        totp_secret: str,
        timeout_seconds: float = REMOTE_REQUEST_TIMEOUT_SECONDS,
        page_size: int = 100,
        transport: httpx.AsyncBaseTransport | None = None,
        remote_session: RemoteAccountSession | None = None,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.username = username
        self.password = password
        self.totp_secret = totp_secret
        self.page_size = page_size
        self._token: str | None = None
        self._remote_session = remote_session
        self._client = httpx.AsyncClient(
            timeout=httpx.Timeout(
                timeout_seconds,
                connect=min(REMOTE_CONNECT_TIMEOUT_SECONDS, timeout_seconds),
            ),
            follow_redirects=False,
            transport=transport,
        )

    async def __aenter__(self) -> RajAdminChargeClient:
        return self

    async def __aexit__(self, *_: object) -> None:
        await self.close()

    async def close(self) -> None:
        await self._client.aclose()

    def _base_headers(self) -> dict[str, str]:
        return {
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh_CN",
            "content-type": "application/json;charset=UTF-8",
            "referer": f"{self.base_url}/",
            "user-agent": "RajDataHandle/0.1",
        }

    async def login(self, *, force: bool = False) -> str:
        if self._remote_session is not None:
            try:
                self._token = await self._remote_session.token(
                    self._login_uncached,
                    force=force,
                    rejected_token=self._token if force else None,
                )
                return self._token
            except RemoteSessionError as exc:
                raise RemoteAuthenticationError(str(exc)) from exc
        if self._token and not force:
            return self._token
        self._token = await self._login_uncached()
        return self._token

    async def _login_uncached(self) -> str:
        try:
            code = generate_totp(self.totp_secret)
        except ValueError as exc:
            raise RemoteAuthenticationError(str(exc)) from exc
        try:
            response = await self._client.post(
                f"{self.base_url}{LOGIN_PATH}",
                headers=self._base_headers(),
                json={
                    "username": self.username,
                    "password": self.password,
                    "code": code,
                },
            )
        except httpx.HTTPError as exc:
            raise RemoteAuthenticationError("远端登录请求失败。") from exc
        if response.status_code == 429:
            raise RemoteAuthenticationError("远端登录次数受限（429）。")
        if response.status_code >= 400:
            raise RemoteAuthenticationError("远端登录返回非成功 HTTP 状态。")
        try:
            payload = response.json()
        except ValueError as exc:
            raise RemoteAuthenticationError("远端登录响应不是有效 JSON。") from exc
        if isinstance(payload, dict) and payload.get("success") is False:
            message = str(payload.get("message") or payload.get("msg") or "").lower()
            if any(word in message for word in ("登录次数", "频繁", "too many", "rate limit")):
                raise RemoteAuthenticationError("远端登录次数受限。")
            raise RemoteAuthenticationError("远端登录被拒绝。")
        token = _extract_token(payload)
        if not token:
            raise RemoteAuthenticationError("远端登录响应中没有 JWT。")
        return token

    async def _get_json(
        self,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        allow_relogin: bool = True,
    ) -> object:
        if path not in REMOTE_GET_PATHS:
            raise RemoteChargeError("GET 请求路径不在远端只读 Allowlist 中。")
        token = await self.login()
        headers = {**self._base_headers(), "authorization": f"Bearer {token}"}
        try:
            response = await self._client.get(
                f"{self.base_url}{path}",
                headers=headers,
                params=params,
            )
        except httpx.HTTPError as exc:
            raise RemoteResponseError("远端只读请求失败。") from exc
        if response_requires_relogin(response) and allow_relogin:
            await self.login(force=True)
            return await self._get_json(path, params=params, allow_relogin=False)
        await self._reject_expired_response(response, token)
        if response.status_code >= 400:
            raise RemoteResponseError("远端只读接口返回非成功 HTTP 状态。")
        try:
            return response.json()
        except ValueError as exc:
            raise RemoteResponseError("远端只读接口响应不是有效 JSON。") from exc

    async def _post_json(
        self,
        path: str,
        *,
        body: dict[str, Any],
        allow_relogin: bool = True,
    ) -> object:
        if path not in REMOTE_POST_PATHS:
            raise RemoteChargeError("POST 请求路径不在远端只读 Allowlist 中。")
        token = await self.login()
        headers = {**self._base_headers(), "authorization": f"Bearer {token}"}
        try:
            response = await self._client.post(
                f"{self.base_url}{path}",
                headers=headers,
                json=body,
            )
        except httpx.HTTPError as exc:
            raise RemoteResponseError("远端只读请求失败。") from exc
        if response_requires_relogin(response) and allow_relogin:
            await self.login(force=True)
            return await self._post_json(path, body=body, allow_relogin=False)
        await self._reject_expired_response(response, token)
        if response.status_code >= 400:
            raise RemoteResponseError("远端只读接口返回非成功 HTTP 状态。")
        try:
            return response.json()
        except ValueError as exc:
            raise RemoteResponseError("远端只读接口响应不是有效 JSON。") from exc

    async def _post_bytes(
        self,
        path: str,
        *,
        body: dict[str, Any],
        allow_relogin: bool = True,
    ) -> bytes:
        if path not in REMOTE_POST_PATHS:
            raise RemoteChargeError("POST 请求路径不在远端只读 Allowlist 中。")
        token = await self.login()
        headers = {**self._base_headers(), "authorization": f"Bearer {token}"}
        try:
            response = await self._client.post(
                f"{self.base_url}{path}",
                headers=headers,
                json=body,
            )
        except httpx.TimeoutException as exc:
            raise RemoteResponseError("远端充值订单导出请求超时。") from exc
        except httpx.HTTPError as exc:
            raise RemoteResponseError("远端充值订单导出请求失败。") from exc
        if response_requires_relogin(response) and allow_relogin:
            await self.login(force=True)
            return await self._post_bytes(path, body=body, allow_relogin=False)
        await self._reject_expired_response(response, token)
        if response.status_code >= 400:
            raise RemoteResponseError("远端充值订单导出返回非成功 HTTP 状态。")
        content = response.content
        if not content.startswith(b"PK"):
            raise RemoteResponseError("远端充值订单导出未返回 Excel 文件。")
        return content

    async def _reject_expired_response(self, response: httpx.Response, token: str) -> None:
        if response_requires_relogin(response):
            if self._remote_session:
                await self._remote_session.reject(token)
            raise RemoteAuthenticationError("重新登录后远端仍拒绝会话，已停止重试。")

    async def fetch_channels(self) -> list[dict[str, str]]:
        data = _response_data(await self._get_json(CHARGE_CHANNEL_PATH))
        if not isinstance(data, list):
            raise RemoteResponseError("远端充值渠道字典结构无效。")
        channels: list[dict[str, str]] = []
        for item in data:
            if not isinstance(item, dict):
                raise RemoteResponseError("远端充值渠道字典包含无效条目。")
            code = str(item.get("value") or "").strip()
            label = str(item.get("label") or "").strip()
            if not code or not label:
                raise RemoteResponseError("远端充值渠道字典缺少代码或名称。")
            channels.append({"code": code, "label": label})
        if not channels:
            raise RemoteResponseError("远端充值渠道名称字典为空。")
        return channels

    async def fetch_payment_channels(self) -> list[dict[str, str]]:
        """Fetch the key/title dictionary whose key is the order pay_method value."""

        data = _response_data(
            await self._get_json(
                DATA_DICTIONARY_PATH,
                params={"code": "pay_channel"},
            )
        )
        if not isinstance(data, list):
            raise RemoteResponseError("远端支付渠道字典结构无效。")

        channels_by_code: dict[str, str] = {}
        for item in data:
            if not isinstance(item, dict):
                raise RemoteResponseError("远端支付渠道字典包含无效条目。")
            code = _text(item.get("key"))
            label = _text(item.get("title"))
            if not code or not label:
                raise RemoteResponseError("远端支付渠道字典缺少 pay_method 值或展示内容。")
            previous = channels_by_code.get(code)
            if previous is not None and previous != label:
                raise RemoteResponseError("远端支付渠道字典中同一 pay_method 对应多个展示内容。")
            channels_by_code[code] = label
        if not channels_by_code:
            raise RemoteResponseError("远端支付渠道字典为空。")
        return [{"code": code, "label": label} for code, label in sorted(channels_by_code.items())]

    def _charge_params(
        self,
        *,
        page: int,
        channel_code: str,
        create_start: str | None = None,
        create_end: str | None = None,
        order_num: str = "",
        out_trade_no: str = "",
    ) -> dict[str, Any]:
        return {
            "page": page,
            "pageSize": self.page_size,
            "create_time[0]": create_start or "",
            "create_time[1]": create_end or "",
            "uid": "",
            "first_pay": "",
            "order_num": order_num,
            "charge_id": "",
            "charge_type": "",
            "pay_channel_type": "",
            "pay_channel_name": "",
            "pay_method": channel_code,
            "pay_type": "",
            "out_trade_no": out_trade_no,
            "status": "",
            "recent": 0,
        }

    async def _fetch_charge_page(
        self,
        *,
        page: int,
        channel_code: str,
        channel_label: str,
        create_start: str | None = None,
        create_end: str | None = None,
        order_num: str = "",
        out_trade_no: str = "",
    ) -> tuple[list[dict[str, Any]], dict[str, int]]:
        payload = await self._get_json(
            CHARGE_ORDER_INDEX_PATH,
            params=self._charge_params(
                page=page,
                channel_code=channel_code,
                create_start=create_start,
                create_end=create_end,
                order_num=order_num,
                out_trade_no=out_trade_no,
            ),
        )
        data = _response_data(payload)
        if not isinstance(data, dict):
            raise RemoteResponseError("远端充值订单 data 结构无效。")
        items = data.get("items")
        page_info = data.get("pageInfo")
        if not isinstance(items, list) or not isinstance(page_info, dict):
            raise RemoteResponseError("远端充值订单分页结构无效。")
        normalized: list[dict[str, Any]] = []
        for item in items:
            if not isinstance(item, dict):
                raise RemoteResponseError("远端充值订单包含无效条目。")
            normalized.append(
                {
                    **item,
                    "_remote_channel_code": channel_code,
                    "_remote_channel_label": channel_label,
                }
            )
        try:
            normalized_page_info = {
                "total": int(page_info.get("total") or 0),
                "current_page": int(page_info.get("currentPage") or page),
                "total_page": int(page_info.get("totalPage") or 0),
            }
        except (TypeError, ValueError) as exc:
            raise RemoteResponseError("远端充值订单分页数字无效。") from exc
        if normalized_page_info["current_page"] != page:
            raise RemoteResponseError("远端充值订单返回页码与请求不一致。")
        return normalized, normalized_page_info

    async def fetch_all_charge_orders(
        self,
        *,
        channels: list[dict[str, str]],
        create_start: str,
        create_end: str,
        on_page_fetched: Callable[[], Awaitable[None]] | None = None,
    ) -> ChargeFetchResult:
        all_orders: list[dict[str, Any]] = []
        fetched_pages = 0
        remote_total = 0
        for channel in channels:
            page = 1
            channel_orders: list[dict[str, Any]] = []
            expected_total = 0
            while True:
                items, page_info = await self._fetch_charge_page(
                    page=page,
                    channel_code=channel["code"],
                    channel_label=channel["label"],
                    create_start=create_start,
                    create_end=create_end,
                )
                fetched_pages += 1
                channel_orders.extend(normalize_charge_order(item) for item in items)
                expected_total = page_info["total"]
                total_page = page_info["total_page"]
                if on_page_fetched is not None:
                    await on_page_fetched()
                if total_page > MAX_CHARGE_PAGES_PER_CHANNEL:
                    raise RemoteResponseError("充值订单数量过多，请缩小查询时间范围。")
                if page >= total_page:
                    break
                page += 1
            if len(channel_orders) != expected_total:
                raise RemoteResponseError("远端充值订单累计数量与 pageInfo.total 不一致。")
            all_orders.extend(channel_orders)
            remote_total += expected_total

        unique_orders: dict[str, dict[str, Any]] = {}
        anonymous_orders: list[dict[str, Any]] = []
        for order in all_orders:
            order_id = str(order.get("id") or "")
            if order_id:
                unique_orders[order_id] = order
            else:
                anonymous_orders.append(order)
        normalized_orders = [*unique_orders.values(), *anonymous_orders]
        return ChargeFetchResult(
            orders=normalized_orders,
            fetched_pages=fetched_pages,
            remote_total=remote_total,
            # Some channel filters can overlap.  De-duplication by remote ID is
            # expected and a complete traversal remains authoritative.
            complete=True,
        )

    @staticmethod
    def _charge_export_body(*, create_start: str, create_end: str) -> dict[str, Any]:
        return {
            "page": 1,
            "pageSize": 10,
            "create_time": [create_start, create_end],
            "uid": "",
            "first_pay": "",
            "channel": [],
            "order_num": "",
            "charge_id": "",
            "charge_type": "",
            "pay_channel_type": "",
            "pay_channel_name": "",
            "pay_method": "",
            "pay_type": "",
            "out_trade_no": "",
            "status": "",
            "update_time": [],
            "recent": 0,
        }

    async def export_charge_orders(
        self,
        *,
        create_start: str,
        create_end: str,
    ) -> ChargeFetchResult:
        """Export one full calendar-day workbook and normalize its approved rows."""

        body = self._charge_export_body(create_start=create_start, create_end=create_end)
        workbook = await self._post_bytes(CHARGE_ORDER_EXPORT_PATH, body=body)
        export_task = {
            **body,
            "status": 1,
            "export_type": 2,
            "operate_type": 3,
            "download": "operate/chargeOrder/export",
        }
        task_data = _response_data(await self._post_json(EXPORT_TASK_SAVE_PATH, body=export_task))
        if not isinstance(task_data, dict) or not _text(task_data.get("id")):
            raise RemoteResponseError("远端充值订单导出任务记录无效。")
        orders = parse_charge_order_export(workbook)
        return ChargeFetchResult(
            orders=orders,
            fetched_pages=1,
            remote_total=len(orders),
            complete=True,
        )

    async def exact_search(
        self,
        *,
        channels: list[dict[str, str]],
        platform_order_no: str | None,
        create_start: str,
        create_end: str,
    ) -> ExactSearchResult:
        found: dict[str, dict[str, Any]] = {}
        complete = True
        if not platform_order_no:
            return ExactSearchResult(orders=[], complete=complete)
        for channel in channels:
            try:
                page = 1
                channel_orders: list[dict[str, Any]] = []
                expected_total = 0
                while True:
                    items, page_info = await self._fetch_charge_page(
                        page=page,
                        channel_code=channel["code"],
                        channel_label=channel["label"],
                        create_start=create_start,
                        create_end=create_end,
                        out_trade_no=platform_order_no,
                    )
                    channel_orders.extend(items)
                    expected_total = page_info["total"]
                    total_page = page_info["total_page"]
                    if total_page > 100_000:
                        raise RemoteResponseError("远端精确复查分页数量异常。")
                    if page >= total_page:
                        break
                    page += 1
                if len(channel_orders) != expected_total:
                    complete = False
                    continue
                for item in channel_orders:
                    identity = str(
                        item.get("id") or item.get("order_num") or item.get("out_trade_no") or ""
                    )
                    if identity:
                        found[identity] = item
            except RemoteChargeError:
                complete = False
        return ExactSearchResult(orders=list(found.values()), complete=complete)
