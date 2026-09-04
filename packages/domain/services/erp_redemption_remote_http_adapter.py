"""Concrete Raj admin-backend adapter for authorised redemption executions.

The class is deliberately not registered as an application dependency and no
HTTP route constructs it.  It can only be used by an execution runner that has
already obtained an :class:`ErpRemoteExecutionGrant`; this keeps migrated API
semantics available without enabling remote writes during code deployment.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any, Literal
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook

from packages.common.totp import generate_totp
from packages.domain.services.erp_redemption_remote_adapter import (
    ErpRedemptionRemoteAdapter,
    RemoteCancelPublishCommand,
    RemoteCancelPublishResult,
    RemoteCreateCommand,
    RemoteCreateResult,
    RemoteDownloadCommand,
    RemoteDownloadResult,
    RemotePublishCommand,
    RemotePublishResult,
)
from packages.domain.services.erp_redemption_remote_gate import ErpRemoteExecutionGrant

LOGIN_PATH = "/api/system/login"
CREATE_PATH = "/api/common/giftCodeConfig/save"
INDEX_PATH = "/api/common/giftCodeConfig/index"
EXPORT_PATH = "/api/common/giftCodeConfig/export"
PUBLISH_PATH = "/api/common/publishTask/save"
CANCEL_PATH = "/api/common/publishTask/cancelAuto"
TAGS_PATH = "/api/common/profileTag/remote"
MAX_EXPORT_BYTES = 16 * 1024 * 1024


class ErpRedemptionRemoteHttpError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ErpRemoteAccountReadGrant:
    account_id: str
    source_id: str
    operation: Literal["CHECK", "TAGS"]
    capability: Literal["ERP_REMOTE_CHECK", "ERP_TAG_READ", "ERP_TAG_SYNC"]


@dataclass(frozen=True, slots=True)
class RemoteDirectoryTag:
    id: int
    name: str


def _clean(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _nested(payload: object, path: str) -> object | None:
    current = payload
    for key in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _first_text(payload: object, *paths: str) -> str | None:
    for path in paths:
        value = _clean(_nested(payload, path))
        if value:
            return value
    return None


def _extract_token(payload: object) -> str | None:
    if isinstance(payload, dict):
        for key in ("token", "jwt", "access_token", "accessToken"):
            if token := _clean(payload.get(key)):
                return token
        for value in payload.values():
            if token := _extract_token(value):
                return token
    elif isinstance(payload, list):
        for value in payload:
            if token := _extract_token(value):
                return token
    return None


def _success_payload(response: httpx.Response) -> object:
    try:
        payload = response.json()
    except ValueError as exc:
        raise ErpRedemptionRemoteHttpError("远端管理后台返回了无法识别的数据。") from exc
    if isinstance(payload, dict):
        code = payload.get("code")
        if payload.get("success") is False or (code is not None and str(code) not in {"0", "200"}):
            raise ErpRedemptionRemoteHttpError(
                _clean(payload.get("message")) or "远端管理后台请求失败。"
            )
    return payload


def _extract_single_code(content: bytes) -> str:
    if not content.startswith(b"PK") or len(content) > MAX_EXPORT_BYTES:
        raise ErpRedemptionRemoteHttpError("远端兑换码文件为空、格式无效或超过限制。")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ErpRedemptionRemoteHttpError("远端兑换码文件缺少表头。")
        headers = [_clean(value) or "" for value in header]
        if "兑换码号码" not in headers:
            raise ErpRedemptionRemoteHttpError("远端兑换码文件缺少兑换码号码列。")
        index = headers.index("兑换码号码")
        codes = {
            code for row in rows if index < len(row) and (code := _clean(row[index])) is not None
        }
    except ErpRedemptionRemoteHttpError:
        raise
    except Exception as exc:
        raise ErpRedemptionRemoteHttpError("远端兑换码文件不是有效 Excel。") from exc
    finally:
        if "workbook" in locals():
            workbook.close()
    if len(codes) != 1:
        raise ErpRedemptionRemoteHttpError("远端兑换码文件必须恰好包含一个兑换码。")
    return next(iter(codes))


class RajAdminGiftCodeAdapter(ErpRedemptionRemoteAdapter):
    """Exact HTTP mapping for the current cloud ERP remote operations."""

    def __init__(
        self,
        *,
        account_id: str,
        source_id: str,
        base_url: str,
        username: str,
        password: str,
        totp_secret: str,
        business_timezone: str,
        timeout_seconds: float = 30,
        transport: httpx.AsyncBaseTransport | None = None,
    ) -> None:
        self.account_id = account_id
        self.source_id = source_id
        self.base_url = base_url.rstrip("/")
        self.username = username
        self.password = password
        self.totp_secret = totp_secret
        self.business_timezone = ZoneInfo(business_timezone)
        self._token: str | None = None
        self._client = httpx.AsyncClient(
            timeout=httpx.Timeout(timeout_seconds, connect=min(10, timeout_seconds)),
            follow_redirects=False,
            transport=transport,
        )

    async def __aenter__(self) -> RajAdminGiftCodeAdapter:
        return self

    async def __aexit__(self, *_: object) -> None:
        await self.close()

    async def close(self) -> None:
        await self._client.aclose()

    def _assert_grant(self, grant: ErpRemoteExecutionGrant, operation: str) -> None:
        if (
            grant.account_id != self.account_id
            or grant.source_id != self.source_id
            or grant.operation != operation
        ):
            raise ErpRedemptionRemoteHttpError("远端执行授权与账号或操作不匹配。")

    def _assert_read_grant(
        self, grant: ErpRemoteAccountReadGrant, operation: Literal["CHECK", "TAGS"]
    ) -> None:
        capability_valid = (
            operation == "CHECK" and grant.capability == "ERP_REMOTE_CHECK"
        ) or (
            operation == "TAGS" and grant.capability in {"ERP_TAG_READ", "ERP_TAG_SYNC"}
        )
        if (
            grant.account_id != self.account_id
            or grant.source_id != self.source_id
            or grant.operation != operation
            or not capability_valid
        ):
            raise ErpRedemptionRemoteHttpError("远端读取授权与账号或操作不匹配。")

    def _headers(self, token: str | None = None) -> dict[str, str]:
        headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh_CN",
            "content-type": "application/json;charset=UTF-8",
            "origin": self.base_url,
            "referer": f"{self.base_url}/",
            "user-agent": "RajDataHandle/0.1",
        }
        headers["authorization"] = f"Bearer {token}" if token else "Bearer null"
        return headers

    async def _login(self, *, force: bool = False) -> str:
        if self._token and not force:
            return self._token
        try:
            response = await self._client.post(
                f"{self.base_url}{LOGIN_PATH}",
                headers=self._headers(),
                json={
                    "username": self.username,
                    "password": self.password,
                    "code": generate_totp(self.totp_secret),
                },
            )
        except (httpx.HTTPError, ValueError) as exc:
            raise ErpRedemptionRemoteHttpError("远端登录失败。") from exc
        if response.status_code >= 400:
            raise ErpRedemptionRemoteHttpError("远端登录返回非成功 HTTP 状态。")
        token = _extract_token(_success_payload(response))
        if not token:
            raise ErpRedemptionRemoteHttpError("远端登录响应中没有访问令牌。")
        self._token = token
        return token

    async def _request(
        self,
        method: str,
        path: str,
        *,
        json: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        bytes_response: bool = False,
        allow_relogin: bool = True,
    ) -> tuple[object | bytes, str | None]:
        token = await self._login()
        try:
            response = await self._client.request(
                method,
                f"{self.base_url}{path}",
                headers=self._headers(token),
                json=json,
                params=params,
            )
        except httpx.HTTPError as exc:
            raise ErpRedemptionRemoteHttpError("无法连接远端管理后台。") from exc
        if response.status_code in {401, 403} and allow_relogin:
            await self._login(force=True)
            return await self._request(
                method,
                path,
                json=json,
                params=params,
                bytes_response=bytes_response,
                allow_relogin=False,
            )
        if response.status_code >= 400:
            raise ErpRedemptionRemoteHttpError("远端管理后台返回非成功 HTTP 状态。")
        request_id = _clean(response.headers.get("x-request-id"))
        return (response.content if bytes_response else _success_payload(response), request_id)

    async def _find_group_key(self, configuration_id: str) -> str | None:
        payload, _ = await self._request(
            "GET",
            INDEX_PATH,
            params={
                "page": 1,
                "pageSize": 100,
                "group_key": "",
                "group_desc": "",
                "valid_type": "",
            },
        )
        items = _nested(payload, "data.items")
        if not isinstance(items, list):
            return None
        for item in items:
            if isinstance(item, dict) and _first_text(item, "id") == configuration_id:
                return _first_text(item, "group_key", "groupKey")
        return None

    async def check_connection(
        self, *, grant: ErpRemoteAccountReadGrant
    ) -> tuple[str, str | None]:
        self._assert_read_grant(grant, "CHECK")
        _, request_id = await self._request(
            "GET",
            INDEX_PATH,
            params={
                "page": 1,
                "pageSize": 1,
                "group_key": "",
                "group_desc": "",
                "valid_type": "",
            },
        )
        return "连接正常，已验证远端兑换码配置访问权限。", request_id

    async def fetch_tags(
        self, *, grant: ErpRemoteAccountReadGrant
    ) -> tuple[list[RemoteDirectoryTag], str | None]:
        self._assert_read_grant(grant, "TAGS")
        payload, request_id = await self._request(
            "POST",
            TAGS_PATH,
            json={
                "openPage": False,
                "remoteOption": {
                    "select": ["tag_id", "name"],
                    "filter": {"tag_type": ["!=", 1], "status": 1},
                },
                "page": 1,
                "pageSize": 500,
            },
        )
        data = _nested(payload, "data")
        if not isinstance(data, list):
            raise ErpRedemptionRemoteHttpError("远端标签接口返回格式无效。")
        tags: list[RemoteDirectoryTag] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            try:
                tag_id = int(item.get("tag_id") or 0)
            except (TypeError, ValueError):
                continue
            name = _clean(item.get("name"))
            if tag_id > 0 and name:
                tags.append(RemoteDirectoryTag(id=tag_id, name=name))
        return tags, request_id

    async def create_configuration(
        self, *, grant: ErpRemoteExecutionGrant, command: RemoteCreateCommand
    ) -> RemoteCreateResult:
        self._assert_grant(grant, "CREATE")
        options = command.options
        all_users = not command.label_ids
        valid_from = command.valid_from or command.claim_date
        valid_to = command.valid_to or command.claim_date
        payload: dict[str, Any] = {
            "flow_times": options.flow_times,
            "status": "2",
            "is_need_check_uuid": int(options.check_uuid),
            "is_need_check_login_ip": int(options.check_login_ip),
            "is_need_check_register_ip": int(options.check_register_ip),
            "user_type": "0" if all_users else "1",
            "is_need_bind_bank_card": int(options.require_bind_bank_card),
            "is_need_bind_phone": int(options.require_bind_phone),
            "remark": command.description,
            "group_desc": command.description,
            "reward_min": str(command.bonus_amount),
            "reward_max": str(command.bonus_max_amount),
            "key_number": str(options.key_number),
            "single_user_limit": options.single_user_limit,
            "single_key_limit": options.single_key_limit,
            "uuid_reward_limit": options.uuid_reward_limit,
            "login_ip_reward_limit": options.login_ip_reward_limit,
            "register_ip_reward_limit": options.register_ip_reward_limit,
            "valid_time": [
                f"{valid_from:%Y-%m-%d} 00:00:00",
                f"{valid_to:%Y-%m-%d} 23:59:59",
            ],
        }
        if not all_users:
            payload["label_array"] = list(command.label_ids)
        for key, value in {
            "activity_recharge": options.activity_recharge,
            "activity_recharge_count": options.activity_recharge_count,
            "activity_id": options.activity_id,
        }.items():
            if value is not None:
                payload[key] = str(value) if key == "activity_recharge" else value
        response, request_id = await self._request("POST", CREATE_PATH, json=payload)
        configuration_id = _first_text(response, "data.id", "id")
        if not configuration_id:
            raise ErpRedemptionRemoteHttpError("远端创建接口未返回配置 ID。")
        return RemoteCreateResult(
            remote_configuration_id=configuration_id,
            remote_group_key=await self._find_group_key(configuration_id),
            remote_request_id=request_id,
        )

    async def publish(
        self, *, grant: ErpRemoteExecutionGrant, command: RemotePublishCommand
    ) -> RemotePublishResult:
        self._assert_grant(grant, "PUBLISH")
        scheduled = command.mode == "SCHEDULED"
        payload: dict[str, Any] = {
            "env": command.publish_environment,
            "publish_type": "2" if scheduled else "1",
            "cfg_type": 19,
        }
        if scheduled:
            if command.scheduled_publish_at is None:
                raise ErpRedemptionRemoteHttpError("定时发布缺少发布时间。")
            local_time = command.scheduled_publish_at.astimezone(self.business_timezone)
            payload["scheduled_time"] = local_time.strftime("%Y-%m-%d %H:%M:%S")
        response, request_id = await self._request("POST", PUBLISH_PATH, json=payload)
        task_id = _first_text(response, "data.id", "id")
        if not task_id:
            raise ErpRedemptionRemoteHttpError("远端发布接口未返回任务 ID。")
        return RemotePublishResult(
            remote_publish_task_id=task_id,
            scheduled_publish_at=command.scheduled_publish_at,
            remote_request_id=request_id,
        )

    async def download(
        self, *, grant: ErpRemoteExecutionGrant, command: RemoteDownloadCommand
    ) -> RemoteDownloadResult:
        self._assert_grant(grant, "DOWNLOAD")
        group_key = command.remote_group_key or await self._find_group_key(
            command.remote_configuration_id
        )
        if not group_key:
            raise ErpRedemptionRemoteHttpError("远端配置尚未生成兑换码组标识。")
        content, request_id = await self._request(
            "POST", EXPORT_PATH, params={"groupKey": group_key}, json={}, bytes_response=True
        )
        if not isinstance(content, bytes):
            raise ErpRedemptionRemoteHttpError("远端兑换码下载响应无效。")
        return RemoteDownloadResult(
            redemption_code=_extract_single_code(content),
            remote_group_key=group_key,
            remote_request_id=request_id,
        )

    async def cancel_publish(
        self, *, grant: ErpRemoteExecutionGrant, command: RemoteCancelPublishCommand
    ) -> RemoteCancelPublishResult:
        self._assert_grant(grant, "CANCEL")
        response, request_id = await self._request(
            "POST", CANCEL_PATH, json={"id": command.remote_publish_task_id}
        )
        result = _nested(response, "data.ret")
        if str(result) != "1":
            raise ErpRedemptionRemoteHttpError("远端未确认撤销定时发布。")
        return RemoteCancelPublishResult(remote_request_id=request_id)
