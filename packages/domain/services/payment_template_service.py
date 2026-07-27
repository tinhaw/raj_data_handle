from __future__ import annotations

import asyncio
import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, BinaryIO

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from packages.domain.models import (
    PaymentChannelBinding,
    PaymentPlatform,
    PaymentTemplateVersion,
)
from packages.domain.schemas.payment_template import (
    PaymentChannelBindingResponse,
    PaymentPlatformResponse,
    PaymentTemplateResponse,
    TemplateDetectionResponse,
)


class TemplateDetectionError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class HeaderCandidate:
    sheet_name: str
    header_row: int
    headers: list[str]


def _normalize_headers(values: list[object]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in values]


def _xlsx_candidates(
    file_object: BinaryIO,
    *,
    header_row: int | None = None,
) -> list[HeaderCandidate]:
    file_object.seek(0)
    workbook = load_workbook(file_object, read_only=True, data_only=True)
    candidates: list[HeaderCandidate] = []
    try:
        for worksheet in workbook.worksheets:
            if header_row is not None:
                row = next(
                    worksheet.iter_rows(
                        min_row=header_row,
                        max_row=header_row,
                        values_only=True,
                    ),
                    None,
                )
                headers = _normalize_headers(list(row or ()))
                if any(headers):
                    candidates.append(
                        HeaderCandidate(
                            sheet_name=worksheet.title,
                            header_row=header_row,
                            headers=headers,
                        )
                    )
                continue
            for row_number, row in enumerate(
                worksheet.iter_rows(max_row=20, values_only=True),
                start=1,
            ):
                headers = _normalize_headers(list(row))
                if any(headers):
                    candidates.append(
                        HeaderCandidate(
                            sheet_name=worksheet.title,
                            header_row=row_number,
                            headers=headers,
                        )
                    )
                    break
    finally:
        workbook.close()
        file_object.seek(0)
    return candidates


def _csv_candidate(
    file_object: BinaryIO,
    *,
    header_row: int | None = None,
) -> list[HeaderCandidate]:
    file_object.seek(0)
    raw = file_object.read(256 * 1024)
    file_object.seek(0)
    decoded: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise TemplateDetectionError("CSV 编码无法识别，请使用 UTF-8 或 GB18030。")
    rows = csv.reader(io.StringIO(decoded))
    for row_number, row in enumerate(rows, start=1):
        headers = _normalize_headers(list(row))
        if header_row is not None:
            if row_number == header_row:
                return (
                    [
                        HeaderCandidate(
                            sheet_name="CSV",
                            header_row=row_number,
                            headers=headers,
                        )
                    ]
                    if any(headers)
                    else []
                )
            continue
        if any(headers):
            return [
                HeaderCandidate(
                    sheet_name="CSV",
                    header_row=row_number,
                    headers=headers,
                )
            ]
        if header_row is None and row_number >= 20:
            break
    return []


async def read_header_candidates(
    upload: UploadFile,
    *,
    header_row: int | None = None,
) -> list[HeaderCandidate]:
    suffix = Path(upload.filename or "").suffix.lower()
    try:
        if suffix == ".xlsx":
            candidates = await asyncio.to_thread(
                _xlsx_candidates,
                upload.file,
                header_row=header_row,
            )
        elif suffix == ".csv":
            candidates = await asyncio.to_thread(
                _csv_candidate,
                upload.file,
                header_row=header_row,
            )
        else:
            raise TemplateDetectionError("当前只支持 .xlsx 和 .csv 文件。")
    except TemplateDetectionError:
        raise
    except Exception as exc:
        raise TemplateDetectionError("文件无法解析，请确认文件未损坏且格式正确。") from exc
    finally:
        await upload.seek(0)
    if not candidates:
        if header_row is not None:
            raise TemplateDetectionError(f"第 {header_row} 行没有可识别的表头字段。")
        raise TemplateDetectionError("文件前 20 行内没有可识别的表头。")
    return candidates


def _template_response(
    template: PaymentTemplateVersion,
    platform: PaymentPlatform,
) -> PaymentTemplateResponse:
    return PaymentTemplateResponse(
        id=template.id,
        platform_id=platform.id,
        platform_key=platform.platform_key,
        platform_display_name=platform.display_name,
        business_type=template.business_type,
        version=template.version,
        sheet_name_pattern=template.sheet_name_pattern,
        header_signature=template.header_signature_json,
        column_mapping=template.column_mapping_json,
        success_status_values=template.success_status_values_json,
        match_rules=template.match_rules_json,
        active=template.active,
    )


async def list_payment_platforms(
    session: AsyncSession,
) -> list[PaymentPlatformResponse]:
    rows = list(
        await session.scalars(select(PaymentPlatform).order_by(PaymentPlatform.platform_key.asc()))
    )
    return [PaymentPlatformResponse.model_validate(row) for row in rows]


async def list_payment_templates(
    session: AsyncSession,
    *,
    business_type: str | None = None,
) -> list[PaymentTemplateResponse]:
    statement = (
        select(PaymentTemplateVersion, PaymentPlatform)
        .join(PaymentPlatform, PaymentPlatform.id == PaymentTemplateVersion.platform_id)
        .order_by(PaymentPlatform.platform_key, PaymentTemplateVersion.version.desc())
    )
    if business_type:
        statement = statement.where(PaymentTemplateVersion.business_type == business_type)
    rows = (await session.execute(statement)).all()
    return [_template_response(template, platform) for template, platform in rows]


async def list_channel_bindings(
    session: AsyncSession,
    *,
    source_id: str | None = None,
    business_type: str | None = None,
) -> list[PaymentChannelBindingResponse]:
    statement = (
        select(PaymentChannelBinding, PaymentPlatform)
        .join(PaymentPlatform, PaymentPlatform.id == PaymentChannelBinding.platform_id)
        .where(PaymentChannelBinding.active.is_(True))
        .order_by(
            PaymentPlatform.platform_key,
            PaymentChannelBinding.remote_channel_code,
        )
    )
    if source_id:
        statement = statement.where(PaymentChannelBinding.source_id == source_id)
    if business_type:
        statement = statement.where(PaymentChannelBinding.business_type == business_type)
    rows = (await session.execute(statement)).all()
    return [
        PaymentChannelBindingResponse(
            id=binding.id,
            platform_id=platform.id,
            platform_key=platform.platform_key,
            source_id=binding.source_id,
            business_type=binding.business_type,
            remote_channel_code=binding.remote_channel_code,
            remote_channel_label=binding.remote_channel_label,
            merchant_discriminator=binding.merchant_discriminator,
            active=binding.active,
        )
        for binding, platform in rows
    ]


async def detect_payment_template(
    session: AsyncSession,
    upload: UploadFile,
    *,
    header_row: int | None = None,
) -> TemplateDetectionResponse:
    candidates = await read_header_candidates(upload, header_row=header_row)
    templates = await list_payment_templates(session)
    best: (
        tuple[
            float,
            int,
            HeaderCandidate,
            PaymentTemplateResponse,
        ]
        | None
    ) = None
    for candidate in candidates:
        detected = {header for header in candidate.headers if header}
        for template in templates:
            expected = {header for header in template.header_signature if header}
            coverage = len(expected & detected) / len(expected) if expected else 0.0
            sheet_matches = (
                not template.sheet_name_pattern
                or re.search(template.sheet_name_pattern, candidate.sheet_name) is not None
            )
            score = (coverage, len(expected))
            if sheet_matches and (best is None or score > (best[0], best[1])):
                best = (coverage, len(expected), candidate, template)

    if best is None or best[0] < 1.0:
        candidate = candidates[0]
        return TemplateDetectionResponse(
            status="unknown",
            file_name=upload.filename or "",
            source_sheet=candidate.sheet_name,
            header_row=candidate.header_row,
            detected_headers=candidate.headers,
            header_coverage=best[0] if best else 0.0,
            template=None,
            message="未识别到已发布模板，请在当前批次确认字段映射。",
        )

    coverage, _, candidate, template = best
    return TemplateDetectionResponse(
        status="matched",
        file_name=upload.filename or "",
        source_sheet=candidate.sheet_name,
        header_row=candidate.header_row,
        detected_headers=candidate.headers,
        header_coverage=coverage,
        template=template,
        message=(
            f"已识别 {template.platform_display_name} "
            f"{template.business_type} 模板 V{template.version}。"
        ),
    )


def detection_snapshot(detection: TemplateDetectionResponse) -> dict[str, Any]:
    return detection.model_dump(by_alias=True, mode="json")
