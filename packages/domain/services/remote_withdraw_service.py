from __future__ import annotations

from collections.abc import Awaitable, Callable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

from packages.domain.services.remote_account_session_service import response_requires_relogin
from packages.domain.services.remote_charge_service import (
    EXPORT_TASK_SAVE_PATH,
    WITHDRAW_ORDER_INDEX_PATH,
    WITHDRAW_STATUS_DICTIONARY_PATH,
    RajAdminChargeClient,
    RemoteResponseError,
    _response_data,
)

MAX_WITHDRAW_PAGES = 200
MAX_WITHDRAW_EXPORT_BYTES = 128 * 1024 * 1024
WITHDRAW_ORDER_EXPORT_PATH = "/api/operate/withdrawOrder/export"

# The remote export contains additional operational and banking columns.  This
# whitelist intentionally names only columns approved for the local analysis
# cache.  The parser never copies unrelated spreadsheet cells into a returned
# row, so account, mobile, bank, IFSC, IP, and free-form reasons remain out of
# the application entirely.
WITHDRAW_EXPORT_COLUMNS = (
    "主键",
    "提现uid",
    "提现订单号",
    "用户渠道",
    "三方支付订单号",
    "支付通道名称",
    "支付通道",
    "提现金额",
    "提现手续费",
    "到账金额",
    "是否首提",
    "状态",
    "创建时间",
    "提交时间",
    "修改时间",
    "审核人",
)
WITHDRAW_EXPORT_STATUS_CODES = {
    "审核拒绝": "-1",
    "待审核": "0",
    "已提交代付": "1",
    "代付失败": "2",
    "代付成功": "3",
    "待审查": "4",
    "提交中": "5",
    "提交三方失败": "6",
}


@dataclass(frozen=True, slots=True)
class WithdrawFetchResult:
    orders: list[dict[str, Any]]
    fetched_pages: int
    remote_total: int
    complete: bool
    # Export paths retain their raw non-empty-row and de-duplication counts so
    # refresh state can explain the difference between an Excel file and the
    # canonical local cache.  Defaults keep the legacy paginated client API
    # fully compatible.
    export_row_count: int = 0
    duplicate_count: int = 0


def _text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized if normalized and normalized != "-" else None


def _amount(value: object) -> str | None:
    normalized = _text(value)
    if normalized is None:
        return None
    try:
        return format(Decimal(normalized), "f")
    except (InvalidOperation, ValueError):
        return None


def _excel_text(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value)


def _export_status(value: object) -> tuple[str, str]:
    """Keep the raw export state until source-dictionary validation runs.

    The worker maps labels to the selected source's dictionary immediately
    before persistence.  Keeping an unfamiliar but non-empty label here lets
    a newly introduced remote status be validated against that dictionary
    instead of hard-coding every possible label in the spreadsheet parser.
    """

    status_label = _excel_text(value) or ""
    if not status_label:
        raise RemoteResponseError("提现订单导出表格包含空状态。")
    if status_label in {"-1", "0", "1", "2", "3", "4", "5", "6"}:
        return status_label, status_label
    return WITHDRAW_EXPORT_STATUS_CODES.get(status_label, status_label), status_label


def normalize_withdraw_order_export_row(item: dict[str, object]) -> dict[str, Any]:
    """Map approved withdrawal-export columns into the local cache schema."""

    status, status_label = _export_status(item.get("状态"))
    return {
        "remote_order_id": _excel_text(item.get("主键")) or "",
        "uid": _excel_text(item.get("提现uid")) or "",
        "order_num": _excel_text(item.get("提现订单号")),
        "channel": _excel_text(item.get("用户渠道")),
        "out_trade_no": _excel_text(item.get("三方支付订单号")),
        "pay_channel_name": _excel_text(item.get("支付通道名称")),
        "pay_channel": _excel_text(item.get("支付通道")),
        "amount": _amount(_excel_text(item.get("提现金额"))),
        "fee": _amount(_excel_text(item.get("提现手续费"))),
        "real_amount": _amount(_excel_text(item.get("到账金额"))),
        "is_first": _excel_text(item.get("是否首提")),
        "status": status,
        "status_label": status_label,
        "create_time": _excel_text(item.get("创建时间")),
        "submit_time": _excel_text(item.get("提交时间")),
        "update_time": _excel_text(item.get("修改时间")),
        "audit_person": _excel_text(item.get("审核人")),
    }


def _parse_withdraw_order_export(
    content: bytes,
) -> tuple[list[dict[str, Any]], int, int]:
    """Parse one Excel export in memory and retain its row-level statistics."""

    if (
        not content
        or len(content) > MAX_WITHDRAW_EXPORT_BYTES
        or not content.startswith(b"PK")
    ):
        raise RemoteResponseError("提现订单导出文件为空、格式无效或超过大小限制。")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception as exc:
        raise RemoteResponseError("远端提现订单导出文件不是有效 Excel 表格。") from exc
    if header_row is None:
        workbook.close()
        raise RemoteResponseError("远端提现订单导出表格缺少表头。")

    headers = [_excel_text(value) or "" for value in header_row]
    missing = [column for column in WITHDRAW_EXPORT_COLUMNS if column not in headers]
    duplicate_headers = [
        column for column in WITHDRAW_EXPORT_COLUMNS if headers.count(column) > 1
    ]
    if missing or duplicate_headers:
        workbook.close()
        raise RemoteResponseError("远端提现订单导出表格表头不符合白名单要求。")

    indexes = {header: index for index, header in enumerate(headers)}
    orders: dict[str, dict[str, Any]] = {}
    row_count = 0
    try:
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                column: values[indexes[column]] if indexes[column] < len(values) else None
                for column in WITHDRAW_EXPORT_COLUMNS
            }
            if not any(value not in (None, "") for value in row.values()):
                continue
            row_count += 1
            order = normalize_withdraw_order_export_row(row)
            if not order["remote_order_id"]:
                raise RemoteResponseError("提现订单导出表格包含缺少主键的行。")
            # The final row for a duplicate remote primary key wins, matching
            # the recharge-export parser and retaining one authoritative order.
            orders[order["remote_order_id"]] = order
    finally:
        workbook.close()
    normalized_orders = list(orders.values())
    return normalized_orders, row_count, row_count - len(normalized_orders)


def parse_withdraw_order_export(content: bytes) -> list[dict[str, Any]]:
    """Read a withdrawal export without persisting the workbook itself."""

    orders, _, _ = _parse_withdraw_order_export(content)
    return orders


def normalize_withdraw_order(item: dict[str, Any]) -> dict[str, Any]:
    """Return only fields approved for the withdrawal monitoring page."""

    time_data = item.get("time")
    if not isinstance(time_data, dict):
        time_data = {}
    return {
        "id": _text(item.get("id")) or "",
        "uid": _text(item.get("uid")) or "",
        "amount": _amount(item.get("amount")),
        "real_amount": _amount(item.get("real_amount")),
        "create_time": _text(time_data.get("create_time") or item.get("create_time")),
        "update_time": _text(time_data.get("update_time") or item.get("update_time")),
        "submit_time": _text(item.get("submit_time") or time_data.get("submit_time")),
        "audit_admin": _text(item.get("audit_admin")),
        "status": _text(item.get("status")) or "",
    }


class RajAdminWithdrawClient(RajAdminChargeClient):
    async def _post_withdraw_export_bytes(
        self,
        *,
        body: dict[str, Any],
        allow_relogin: bool = True,
    ) -> bytes:
        """Call the single allowlisted withdrawal export endpoint.

        This is intentionally separate from the inherited generic POST helper:
        the base client's allowlist covers the legacy pagination endpoint while
        this method has no caller-controlled path at all.
        """

        token = await self.login()
        headers = {**self._base_headers(), "authorization": f"Bearer {token}"}
        try:
            response = await self._client.post(
                f"{self.base_url}{WITHDRAW_ORDER_EXPORT_PATH}",
                headers=headers,
                json=body,
            )
        except httpx.TimeoutException as exc:
            raise RemoteResponseError("远端提现订单导出请求超时。") from exc
        except httpx.HTTPError as exc:
            raise RemoteResponseError("远端提现订单导出请求失败。") from exc
        if response_requires_relogin(response) and allow_relogin:
            await self.login(force=True)
            return await self._post_withdraw_export_bytes(
                body=body,
                allow_relogin=False,
            )
        await self._reject_expired_response(response, token)
        if response.status_code >= 400:
            raise RemoteResponseError("远端提现订单导出返回非成功 HTTP 状态。")
        content = response.content
        if not content.startswith(b"PK"):
            raise RemoteResponseError("远端提现订单导出未返回 Excel 文件。")
        if len(content) > MAX_WITHDRAW_EXPORT_BYTES:
            raise RemoteResponseError("远端提现订单导出文件超过大小限制。")
        return content

    async def fetch_withdraw_statuses(self) -> list[dict[str, str]]:
        """Fetch the fixed remote withdrawal-status dictionary through the read allowlist."""

        data = _response_data(
            await self._get_json(
                WITHDRAW_STATUS_DICTIONARY_PATH,
                params={"code": "withdraw_status"},
            )
        )
        if not isinstance(data, list):
            raise RemoteResponseError("远端提现状态字典结构无效。")

        statuses_by_code: dict[str, str] = {}
        for item in data:
            if not isinstance(item, dict):
                raise RemoteResponseError("远端提现状态字典包含无效条目。")
            code = _text(item.get("key"))
            label = _text(item.get("title"))
            if not code or not label:
                raise RemoteResponseError("远端提现状态字典缺少状态值或展示文案。")
            previous = statuses_by_code.get(code)
            if previous is not None and previous != label:
                raise RemoteResponseError("远端提现状态字典中同一状态值对应多个文案。")
            statuses_by_code[code] = label
        if not statuses_by_code:
            raise RemoteResponseError("远端提现状态字典为空。")
        return [
            {"code": code, "label": label}
            for code, label in sorted(statuses_by_code.items())
        ]

    def _withdraw_body(
        self,
        *,
        page: int,
        create_start: str,
        create_end: str,
        uid: str = "",
        status: str = "",
    ) -> dict[str, Any]:
        return {
            "page": page,
            "pageSize": self.page_size,
            "create_time": [create_start, create_end],
            "uid": uid,
            "channel": [],
            "pay_channel_name": "",
            "pay_channel": "",
            "order_num": "",
            "out_trade_no": "",
            "is_first": "",
            "update_time": [],
            "status": status,
            "not_to_back_cash": "",
        }

    @staticmethod
    def _withdraw_export_body(*, create_start: str, create_end: str) -> dict[str, Any]:
        return {
            "page": 1,
            "pageSize": 10,
            "create_time": [create_start, create_end],
            "uid": "",
            "channel": [],
            "pay_channel_name": "",
            "pay_channel": "",
            "order_num": "",
            "out_trade_no": "",
            "is_first": "",
            "update_time": [],
            "status": "",
            "not_to_back_cash": "",
            "recent": 0,
        }

    async def export_withdraw_orders(
        self,
        *,
        create_start: str,
        create_end: str,
    ) -> WithdrawFetchResult:
        """Export one full calendar-day workbook and normalize approved rows."""

        body = self._withdraw_export_body(create_start=create_start, create_end=create_end)
        workbook = await self._post_withdraw_export_bytes(body=body)
        export_task = {
            **body,
            "status": 1,
            "export_type": 2,
            "operate_type": 3,
            "download": "operate/withdrawOrder/export",
        }
        task_data = _response_data(await self._post_json(EXPORT_TASK_SAVE_PATH, body=export_task))
        if not isinstance(task_data, dict) or not _text(task_data.get("id")):
            raise RemoteResponseError("远端提现订单导出任务记录无效。")
        orders, export_row_count, duplicate_count = _parse_withdraw_order_export(workbook)
        return WithdrawFetchResult(
            orders=orders,
            fetched_pages=1,
            remote_total=len(orders),
            complete=True,
            export_row_count=export_row_count,
            duplicate_count=duplicate_count,
        )

    async def _fetch_withdraw_page(
        self,
        *,
        page: int,
        create_start: str,
        create_end: str,
        uid: str = "",
        status: str = "",
    ) -> tuple[list[dict[str, Any]], dict[str, int]]:
        payload = await self._post_json(
            WITHDRAW_ORDER_INDEX_PATH,
            body=self._withdraw_body(
                page=page,
                create_start=create_start,
                create_end=create_end,
                uid=uid,
                status=status,
            ),
        )
        data = _response_data(payload)
        if not isinstance(data, dict):
            raise RemoteResponseError("远端提现订单 data 结构无效。")
        items = data.get("items")
        page_info = data.get("pageInfo")
        if not isinstance(items, list) or not isinstance(page_info, dict):
            raise RemoteResponseError("远端提现订单分页结构无效。")
        normalized: list[dict[str, Any]] = []
        for item in items:
            if not isinstance(item, dict):
                raise RemoteResponseError("远端提现订单包含无效条目。")
            normalized.append(normalize_withdraw_order(item))
        try:
            normalized_page_info = {
                "total": int(page_info.get("total") or 0),
                "current_page": int(page_info.get("currentPage") or page),
                "total_page": int(page_info.get("totalPage") or 0),
            }
        except (TypeError, ValueError) as exc:
            raise RemoteResponseError("远端提现订单分页数字无效。") from exc
        if normalized_page_info["current_page"] != page:
            raise RemoteResponseError("远端提现订单返回页码与请求不一致。")
        return normalized, normalized_page_info

    async def fetch_all_withdraw_orders(
        self,
        *,
        create_start: str,
        create_end: str,
        uid: str = "",
        status: str = "",
        on_page_fetched: Callable[[], Awaitable[None]] | None = None,
    ) -> WithdrawFetchResult:
        orders: list[dict[str, Any]] = []
        fetched_pages = 0
        remote_total = 0
        page = 1
        while True:
            items, page_info = await self._fetch_withdraw_page(
                page=page,
                create_start=create_start,
                create_end=create_end,
                uid=uid,
                status=status,
            )
            fetched_pages += 1
            orders.extend(items)
            remote_total = page_info["total"]
            total_page = page_info["total_page"]
            if on_page_fetched is not None:
                await on_page_fetched()
            if total_page > MAX_WITHDRAW_PAGES:
                raise RemoteResponseError("提现订单数量过多，请缩小查询时间范围。")
            if page >= total_page:
                break
            page += 1

        unique_orders: dict[str, dict[str, Any]] = {}
        anonymous_orders: list[dict[str, Any]] = []
        for order in orders:
            order_id = str(order.get("id") or "")
            if order_id:
                unique_orders[order_id] = order
            else:
                anonymous_orders.append(order)
        normalized_orders = [*unique_orders.values(), *anonymous_orders]
        return WithdrawFetchResult(
            orders=normalized_orders,
            fetched_pages=fetched_pages,
            remote_total=remote_total,
            complete=len(normalized_orders) == remote_total,
        )
