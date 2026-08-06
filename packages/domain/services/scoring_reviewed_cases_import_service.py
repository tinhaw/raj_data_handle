"""Strict parser for the scoring-review reviewed-cases Excel export.

The scoring-review export is supplementary data only.  Its ``案件号`` is a
foreign key to the withdrawal-export ``主键``; parsing it must never create a
withdrawal order on its own.  The persistence layer is consequently expected
to join ``withdraw_order_id`` to an already-cached withdrawal order and ignore
unmatched cases.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Final
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

MAX_SCORING_REVIEWED_CASES_EXPORT_BYTES: Final = 32 * 1024 * 1024
MAX_SCORING_REVIEWED_CASES_EXPORT_UNCOMPRESSED_BYTES: Final = 128 * 1024 * 1024
MAX_SCORING_REVIEWED_CASES_EXPORT_ROWS: Final = 100_000
MAX_SCORING_REVIEWED_CASES_EXPORT_COLUMNS: Final = 64

# The four source-owned columns are checked as part of the export contract but
# deliberately not copied into ``ScoringReviewedCase``.  Raj Data Handle keeps
# the remote withdrawal export authoritative for UID, amount, channel and
# withdrawal time.
SCORING_REVIEWED_CASES_EXPORT_COLUMNS: Final = (
    "案件号",
    "UID",
    "提现金额",
    "渠道",
    "全局硬性条件",
    "场景审核",
    "评分审核",
    "决断阶段",
    "最终审核建议",
    "操作结果",
    "摘要",
    "当前状态",
    "审核完成时间",
    "审核耗时",
    "队列中耗时",
    "进入队列时间",
    "退出队列时间",
    "提现时间",
)


class ScoringReviewedCasesImportError(ValueError):
    """Safe, user-facing failure raised for an invalid scoring Excel export."""


@dataclass(frozen=True, slots=True)
class ScoringReviewedCase:
    """Whitelisted scoring fields that can enrich an existing withdrawal order."""

    withdraw_order_id: str
    global_hard_condition: str | None
    scenario_review: str | None
    score_review: str | None
    decision_stage: str | None
    final_review_suggestion: str | None
    operation_result: str | None
    review_summary: str | None
    current_status: str | None
    review_completed_at: str | None
    review_duration: str | None
    queue_duration: str | None
    entered_queue_at: str | None
    exited_queue_at: str | None


@dataclass(frozen=True, slots=True)
class ScoringReviewedCasesImportResult:
    """Rows available to join, without any withdrawal-order upsert semantics."""

    cases: list[ScoringReviewedCase]
    source_row_count: int


def _excel_text(value: object, *, limit: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        normalized = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = str(value).strip()
    if not normalized or normalized == "-":
        return None
    if len(normalized) > limit:
        raise ScoringReviewedCasesImportError("评分审核导出表格包含超长字段。")
    return normalized


def _row_is_empty(values: tuple[object, ...]) -> bool:
    return not any(_excel_text(value, limit=8_192) is not None for value in values)


def _validate_xlsx_container(content: bytes) -> None:
    if (
        not content
        or len(content) > MAX_SCORING_REVIEWED_CASES_EXPORT_BYTES
        or not content.startswith(b"PK")
    ):
        raise ScoringReviewedCasesImportError(
            "评分审核导出文件为空、格式无效或超过大小限制。"
        )
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            uncompressed_size = sum(item.file_size for item in entries)
            names = {item.filename for item in entries}
    except (BadZipFile, OSError) as exc:
        raise ScoringReviewedCasesImportError("评分审核导出文件不是有效 Excel 表格。") from exc
    if (
        not entries
        or len(entries) > 2_000
        or uncompressed_size > MAX_SCORING_REVIEWED_CASES_EXPORT_UNCOMPRESSED_BYTES
        or "[Content_Types].xml" not in names
        or "xl/workbook.xml" not in names
    ):
        raise ScoringReviewedCasesImportError("评分审核导出文件不是有效 Excel 表格。")


def _open_worksheet(content: bytes):  # type: ignore[no-untyped-def]
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception as exc:
        raise ScoringReviewedCasesImportError("评分审核导出文件不是有效 Excel 表格。") from exc
    if header_row is None:
        workbook.close()
        raise ScoringReviewedCasesImportError("评分审核导出表格缺少表头。")
    return workbook, worksheet, header_row


def _header_indexes(header_row: tuple[object, ...]) -> dict[str, int]:
    headers = [_excel_text(value, limit=256) or "" for value in header_row]
    missing = [column for column in SCORING_REVIEWED_CASES_EXPORT_COLUMNS if column not in headers]
    duplicate_headers = [
        column for column in SCORING_REVIEWED_CASES_EXPORT_COLUMNS if headers.count(column) > 1
    ]
    if missing or duplicate_headers:
        raise ScoringReviewedCasesImportError("评分审核导出表格表头不符合白名单要求。")
    return {column: headers.index(column) for column in SCORING_REVIEWED_CASES_EXPORT_COLUMNS}


def _value(
    values: tuple[object, ...],
    indexes: dict[str, int],
    column: str,
    *,
    limit: int = 256,
) -> str | None:
    index = indexes[column]
    value = values[index] if index < len(values) else None
    return _excel_text(value, limit=limit)


def parse_scoring_reviewed_cases_export(content: bytes) -> ScoringReviewedCasesImportResult:
    """Parse one scoring-review Excel export into safe join candidates.

    The parser accepts only the known ``.xlsx`` schema and rejects duplicate
    or missing case identifiers.  A caller must subsequently join every
    returned candidate to an existing withdrawal order by
    ``withdraw_order_id``; unmatched candidates must be ignored rather than
    inserted as new withdrawal orders.
    """

    _validate_xlsx_container(content)
    workbook, worksheet, header_row = _open_worksheet(content)
    try:
        if (
            worksheet.max_row - 1 > MAX_SCORING_REVIEWED_CASES_EXPORT_ROWS
            or worksheet.max_column > MAX_SCORING_REVIEWED_CASES_EXPORT_COLUMNS
        ):
            raise ScoringReviewedCasesImportError("评分审核导出文件超过行列数量限制。")
        indexes = _header_indexes(header_row)
        cases: list[ScoringReviewedCase] = []
        seen_case_ids: set[str] = set()
        source_row_count = 0
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if _row_is_empty(values):
                continue
            source_row_count += 1
            withdraw_order_id = _value(values, indexes, "案件号", limit=64)
            if withdraw_order_id is None:
                raise ScoringReviewedCasesImportError("评分审核导出表格包含缺少案件号的行。")
            if withdraw_order_id in seen_case_ids:
                raise ScoringReviewedCasesImportError("评分审核导出表格包含重复案件号。")
            seen_case_ids.add(withdraw_order_id)
            cases.append(
                ScoringReviewedCase(
                    withdraw_order_id=withdraw_order_id,
                    global_hard_condition=_value(values, indexes, "全局硬性条件"),
                    scenario_review=_value(values, indexes, "场景审核"),
                    score_review=_value(values, indexes, "评分审核"),
                    decision_stage=_value(values, indexes, "决断阶段"),
                    final_review_suggestion=_value(values, indexes, "最终审核建议"),
                    operation_result=_value(values, indexes, "操作结果"),
                    review_summary=_value(values, indexes, "摘要", limit=2_000),
                    current_status=_value(values, indexes, "当前状态"),
                    review_completed_at=_value(values, indexes, "审核完成时间"),
                    review_duration=_value(values, indexes, "审核耗时"),
                    queue_duration=_value(values, indexes, "队列中耗时"),
                    entered_queue_at=_value(values, indexes, "进入队列时间"),
                    exited_queue_at=_value(values, indexes, "退出队列时间"),
                )
            )
    finally:
        workbook.close()
    return ScoringReviewedCasesImportResult(cases=cases, source_row_count=source_row_count)
